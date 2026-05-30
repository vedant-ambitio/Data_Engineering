#!/usr/bin/env python3
"""
generate_review_sheet_masters.py — Excel review sheet for Masters low confidence courses
==========================================================================================

Sheet 1 (Review): Key fields extracted via regex from markdown text
Sheet 2 (Full Markdown): Complete markdown content per course

Does NOT modify any existing files. No Gemini API calls — pure regex extraction.
"""

import json
import os
import re

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── Config ──────────────────────────────────────────────────────────────────

EXTRACTED_DIR = "university_data/structured_extraction_test"  # masters extraction
MD_DIR = "classification_results/masters/low_confidence"
OUTPUT_FILE = "masters_review_low_v7.xlsx"


def _strip_citations(md):
    """Remove <citation>...</citation> blocks for cleaner regex matching."""
    return re.sub(r'<citation>.*?</citation>', '', md, flags=re.DOTALL)


def _section(md, header_keywords):
    """Find a section by header keyword and return its text until the next ## header."""
    for kw in header_keywords:
        # Match ## Header containing keyword (case insensitive)
        pattern = rf'##\s+[^\n]*{re.escape(kw)}[^\n]*\n(.*?)(?=\n##\s|\Z)'
        m = re.search(pattern, md, re.IGNORECASE | re.DOTALL)
        if m:
            return m.group(1)
    return ""


def extract_tuition(md):
    """Find tuition fees in markdown."""
    section = _section(md, ["Tuition", "Fees", "Cost"])
    if not section:
        section = md  # search whole doc as fallback

    # Patterns: $XX,XXX, £XX,XXX, €XX,XXX, USD XX,XXX, etc.
    patterns = [
        r'(?:international\s+students?|out[-\s]?of[-\s]?state)[^\n]*?[\$£€][\s]?(\d{1,3}(?:,\d{3})*(?:\.\d+)?)',
        r'[\$£€][\s]?(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(?:per\s+year|/year|annually|annual)',
        r'(?:tuition|fee)[^\n]*?[\$£€][\s]?(\d{1,3}(?:,\d{3})*(?:\.\d+)?)',
        r'(?:USD|GBP|EUR|CAD|AUD)\s*(\d{1,3}(?:,\d{3})*(?:\.\d+)?)',
    ]
    found = []
    for pat in patterns:
        for m in re.finditer(pat, section, re.IGNORECASE):
            try:
                val = float(m.group(1).replace(",", ""))
                if 1000 < val < 200000:  # sanity check
                    found.append((val, m.group(0)[:80]))
            except ValueError:
                continue
        if found:
            break

    if found:
        # Pick largest (international tuition is usually highest)
        found.sort(key=lambda x: -x[0])
        val, ctx = found[0]
        # Detect currency
        if "£" in ctx or "GBP" in ctx.upper():
            return f"GBP {val:,.0f}/year"
        elif "€" in ctx or "EUR" in ctx.upper():
            return f"EUR {val:,.0f}/year"
        elif "AUD" in ctx.upper():
            return f"AUD {val:,.0f}/year"
        elif "CAD" in ctx.upper():
            return f"CAD {val:,.0f}/year"
        else:
            return f"USD {val:,.0f}/year"
    return None


def extract_deadlines_by_intake(md):
    """
    Extract deadlines split by intake. Returns dict like:
    {"fall_2026": "...", "spring_2027": "...", "fall_2027": "...", "other": "..."}
    """
    section = _section(md, ["Application Deadline", "Deadline", "Intake", "Important Dates"])
    if not section:
        return {}

    date_kw = ["deadline", "round", "decision", "priority", "application", "submit", "due"]
    date_patterns = [
        r'(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2}(?:,\s*\d{4})?',
        r'\d{1,2}\s+(January|February|March|April|May|June|July|August|September|October|November|December)(?:\s+\d{4})?',
        r'\d{4}-\d{2}-\d{2}',
        r'\d{1,2}/\d{1,2}/\d{2,4}',
    ]

    # Split section into intake blocks: bullet line with "Intake" or month + year header
    intake_pattern = re.compile(
        r'\*?\*?\s*\*?\s*(Fall|Spring|Summer|Winter|Autumn|January|February|March|April|May|June|July|August|September|October|November|December|Semester)[^\n:]*?(20\d{2})[^\n]*?:?',
        re.IGNORECASE,
    )

    # Find all intake headers and their positions
    intake_blocks = []
    matches = list(intake_pattern.finditer(section))
    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(section)
        header = m.group(0).strip("* :").strip()
        body = section[start:end]

        # Detect season + year
        text = header.lower()
        year_m = re.search(r'20\d{2}', header)
        year = year_m.group(0) if year_m else ""

        if "fall" in text or "autumn" in text or "semester 1" in text or "september" in text or "august" in text:
            season = "fall"
        elif "spring" in text or "semester 2" in text or "january" in text or "february" in text:
            season = "spring"
        elif "summer" in text or "june" in text or "july" in text or "may" in text:
            season = "summer"
        elif "winter" in text or "december" in text:
            season = "winter"
        else:
            season = "other"

        key = f"{season}_{year}" if year else season
        intake_blocks.append((key, body))

    # For each intake, extract deadline lines
    result = {}
    for key, body in intake_blocks:
        lines_with_dates = []
        for line in body.split("\n"):
            line = line.strip().lstrip("*-• ").strip()
            if not line:
                continue
            if any(kw in line.lower() for kw in date_kw) or any(re.search(p, line, re.IGNORECASE) for p in date_patterns):
                if any(re.search(p, line, re.IGNORECASE) for p in date_patterns):
                    clean = line.replace("**", "").strip()
                    if len(clean) < 250 and "information not available" not in clean.lower():
                        lines_with_dates.append(clean)
        if lines_with_dates:
            existing = result.get(key, "")
            combined = "; ".join(lines_with_dates[:3])
            result[key] = f"{existing}; {combined}" if existing else combined

    return result


def extract_deadlines(md):
    """Backward compatible: returns combined string."""
    parts = extract_deadlines_by_intake(md)
    if not parts:
        return None
    return "; ".join(f"[{k}] {v}" for k, v in parts.items())


def extract_test_score(md, test_name):
    """Extract minimum score for TOEFL/IELTS/Duolingo/PTE."""
    section = _section(md, ["English Proficiency", "Admission Requirements", "Test", "Standardized"])
    if not section:
        section = md

    if test_name.upper() == "TOEFL":
        patterns = [
            r'TOEFL[^.\n]*?(?:overall\s+score\s+of\s+|score[:\s]+|minimum[:\s]+|requires?\s+|of\s+)(\d{2,3})',
            r'TOEFL[^.\n]*?(\d{2,3})',
        ]
    elif test_name.upper() == "IELTS":
        patterns = [
            r'IELTS[^.\n]*?(?:overall\s+score\s+of\s+|score[:\s]+|minimum[:\s]+|of\s+)(\d\.?\d?)',
            r'IELTS[^.\n]*?(\d\.\d|\d\.0)',
        ]
    elif test_name.upper() == "DUOLINGO":
        patterns = [
            r'Duolingo[^.\n]*?(?:overall\s+score\s+of\s+|score[:\s]+|minimum[:\s]+|of\s+)(\d{2,3})',
            r'(?:DET|Duolingo)[^.\n]*?(\d{2,3})',
        ]
    elif test_name.upper() == "PTE":
        patterns = [
            r'PTE[^.\n]*?(?:overall\s+score\s+of\s+|score[:\s]+|minimum[:\s]+|of\s+)(\d{2,3})',
            r'PTE[^.\n]*?(\d{2,3})',
        ]
    else:
        return None

    for pat in patterns:
        m = re.search(pat, section, re.IGNORECASE)
        if m:
            val = m.group(1)
            # Sanity checks
            try:
                num = float(val)
                if test_name.upper() == "TOEFL" and not (40 <= num <= 120):
                    continue
                if test_name.upper() == "DUOLINGO" and not (60 <= num <= 160):
                    continue
                if test_name.upper() == "PTE" and not (30 <= num <= 90):
                    continue
            except ValueError:
                continue
            return val
    return None


def extract_gre_score(md):
    """Extract GRE score requirement (min or avg)."""
    section = _section(md, ["GRE", "Standardized Tests", "Admission Requirements", "Test"])
    if not section:
        section = md

    # Look for "GRE ... XXX" patterns
    patterns = [
        r'GRE[^.\n]*?(?:minimum|min|of|score|requires?)[:\s]+(\d{3})',
        r'GRE[^.\n]*?(\d{3})\s*(?:minimum|or higher|required)',
        r'(?:Verbal|Quant(?:itative)?)[^.\n]*?(\d{3})\s*(?:and|,)\s*(?:Verbal|Quant)[^.\n]*?(\d{3})',
        r'GRE.*?Verbal[:\s]+(\d{3})[^.\n]*?Quant[^.\n]*?(\d{3})',
    ]
    for pat in patterns:
        m = re.search(pat, section, re.IGNORECASE)
        if m:
            if len(m.groups()) == 2:
                return f"V:{m.group(1)} Q:{m.group(2)}"
            val = m.group(1)
            try:
                num = int(val)
                if 130 <= num <= 170:
                    return f"Section: {val}"
                if 260 <= num <= 340:
                    return f"Total: {val}"
            except ValueError:
                continue
    return None


def extract_lor_count(md):
    """Find number of letters of recommendation required."""
    section = _section(md, ["Letters of Recommendation", "Recommendations", "Admission Requirements", "Required documents"])
    if not section:
        section = md

    patterns = [
        r'(\d+|one|two|three|four|five)\s+(?:letters?\s+of\s+recommendation|recommendation\s+letters?|references?)',
        r'(?:letters?\s+of\s+recommendation|recommendation\s+letters?|references?)[^.\n]*?(\d+|one|two|three|four|five)',
        r'(\d+)\s*LORs?',
    ]
    word_to_num = {"one": "1", "two": "2", "three": "3", "four": "4", "five": "5"}
    for pat in patterns:
        m = re.search(pat, section, re.IGNORECASE)
        if m:
            val = m.group(1).lower()
            return word_to_num.get(val, val)
    return None


def extract_app_fee(md):
    """Find application fee."""
    section = _section(md, ["Application Fee", "Application", "Admission Requirements", "Fees"])
    if not section:
        section = md

    # Patterns for app fee
    patterns = [
        r'application\s+fee[^.\n]*?[\$£€][\s]?(\d{1,4}(?:\.\d{2})?)',
        r'application\s+fee[^.\n]*?(?:USD|GBP|EUR|CAD|AUD)\s*(\d{1,4}(?:\.\d{2})?)',
        r'application\s+fee[^.\n]*?(?:is|of|:)\s*[\$£€]?\s*(\d{1,4}(?:\.\d{2})?)',
        r'(?:non[-\s]?refundable\s+)?fee[^.\n]*?[\$£€][\s]?(\d{1,4}(?:\.\d{2})?)',
    ]
    for pat in patterns:
        m = re.search(pat, section, re.IGNORECASE)
        if m:
            try:
                val = float(m.group(1))
                if 10 <= val <= 500:  # sanity check
                    ctx = m.group(0)
                    if "£" in ctx or "GBP" in ctx.upper():
                        return f"GBP {val:.0f}"
                    elif "€" in ctx or "EUR" in ctx.upper():
                        return f"EUR {val:.0f}"
                    elif "AUD" in ctx.upper():
                        return f"AUD {val:.0f}"
                    elif "CAD" in ctx.upper():
                        return f"CAD {val:.0f}"
                    return f"USD {val:.0f}"
            except ValueError:
                continue
    return None


def extract_test_required(md, test_name):
    """Detect if GRE/GMAT is required, optional, waived, or not mentioned."""
    section = _section(md, ["Admission Requirements", "Standardized Tests", "Test", "Application"])
    if not section:
        section = md

    # Look for the test name and what's around it
    pattern = rf'{test_name}[^.\n]{{0,200}}'
    matches = re.findall(pattern, section, re.IGNORECASE)
    if not matches:
        return None

    text = " ".join(matches).lower()

    if "not required" in text or "no longer required" in text or "waived" in text:
        return "Not required / Waived"
    if "optional" in text or "test-optional" in text:
        return "Optional"
    if "required" in text or "must submit" in text or "mandatory" in text:
        return "Required"
    if "information not available" in text or "not available" in text or "not specified" in text:
        return None
    if "recommended" in text or "encouraged" in text:
        return "Recommended"
    return "Mentioned (check details)"


def extract_work_experience(md):
    """Find work experience requirements."""
    section = _section(md, ["Work Experience", "Admission Requirements", "Eligibility"])
    if not section:
        section = md

    # Look for "X years" near "work experience" or "professional experience"
    patterns = [
        r'(\d+)\s*[-+]?\s*(?:to\s+\d+\s+)?years?\s+(?:of\s+)?(?:professional\s+|relevant\s+|work\s+)experience',
        r'(?:work|professional)\s+experience[^.\n]*?(\d+)\s*[-+]?\s*years?',
        r'minimum\s+of\s+(\d+)\s*years?\s+(?:work|professional)',
    ]
    for pat in patterns:
        m = re.search(pat, section, re.IGNORECASE)
        if m:
            return f"{m.group(1)}+ years"

    # Check for "not required" / "preferred"
    we_text = re.search(r'work\s+experience[^.\n]{0,150}', section, re.IGNORECASE)
    if we_text:
        text = we_text.group(0).lower()
        if "not required" in text or "not explicitly" in text:
            return "Not required"
        if "preferred" in text or "recommended" in text:
            return "Preferred (no specific years)"
    return None


def extract_gpa(md):
    """Find minimum GPA requirement."""
    section = _section(md, ["GPA", "Admission Requirements", "Eligibility"])
    if not section:
        section = md

    patterns = [
        r'GPA[^.\n]*?(\d\.\d{1,2})\s*/\s*(\d\.\d?)',
        r'(\d\.\d{1,2})\s*/\s*(\d\.\d?)\s*GPA',
        r'minimum\s+GPA[^.\n]*?(\d\.\d{1,2})',
        r'GPA\s+of\s+(\d\.\d{1,2})',
        r'(\d:\d)\s*(?:Bachelors?|Honours?|degree)',  # UK degree class like 2:2
    ]
    for pat in patterns:
        m = re.search(pat, section, re.IGNORECASE)
        if m:
            if len(m.groups()) == 2 and m.group(2):
                return f"{m.group(1)}/{m.group(2)}"
            return m.group(1)
    return None


def extract_official_link(md):
    """Find the first university course/program URL in citations."""
    # Look in citation blocks for URLs that look like program pages
    urls = re.findall(r'https?://[^\s<>"\)]+', md)
    # Prefer URLs containing program-related keywords
    priority = ["postgraduate", "graduate", "program", "course", "msc", "ma/", "mba", "admission"]
    for url in urls:
        url_lower = url.lower()
        if any(kw in url_lower for kw in priority):
            return url.rstrip(".,;)")
    # Fallback: first URL
    if urls:
        return urls[0].rstrip(".,;)")
    return None


def extract_fields_from_md(md):
    """Extract all key fields directly from markdown text via regex."""
    if not md:
        return {}

    md_clean = _strip_citations(md)
    deadline_parts = extract_deadlines_by_intake(md_clean)

    # Bucket deadlines into common columns
    fall_2026 = deadline_parts.get("fall_2026", "")
    spring_2027 = deadline_parts.get("spring_2027", "")
    fall_2027 = deadline_parts.get("fall_2027", "")
    summer_2026 = deadline_parts.get("summer_2026", "")
    summer_2027 = deadline_parts.get("summer_2027", "")
    winter_2026 = deadline_parts.get("winter_2026", "")
    winter_2027 = deadline_parts.get("winter_2027", "")

    # Anything else goes to "other"
    used_keys = {"fall_2026", "spring_2027", "fall_2027", "summer_2026", "summer_2027",
                 "winter_2026", "winter_2027"}
    other_parts = [f"[{k}] {v}" for k, v in deadline_parts.items() if k not in used_keys]
    other = "; ".join(other_parts) if other_parts else ""

    return {
        "deadline_fall_2026": fall_2026,
        "deadline_spring_2027": spring_2027,
        "deadline_fall_2027": fall_2027,
        "deadline_summer_2026": summer_2026,
        "deadline_summer_2027": summer_2027,
        "deadline_winter_2026": winter_2026,
        "deadline_winter_2027": winter_2027,
        "deadline_other": other,
        "tuition_fees": extract_tuition(md_clean),
        "application_fee": extract_app_fee(md_clean),
        "TOEFL_min": extract_test_score(md_clean, "TOEFL"),
        "IELTS_min": extract_test_score(md_clean, "IELTS"),
        "Duolingo_min": extract_test_score(md_clean, "DUOLINGO"),
        "PTE_min": extract_test_score(md_clean, "PTE"),
        "GRE_required": extract_test_required(md_clean, "GRE"),
        "GRE_score": extract_gre_score(md_clean),
        "GMAT_required": extract_test_required(md_clean, "GMAT"),
        "LOR_count": extract_lor_count(md_clean),
        "work_experience": extract_work_experience(md_clean),
        "GPA_min": extract_gpa(md_clean),
        "officialPageLink": extract_official_link(md),
    }


# Keep the old function name as alias for compatibility
def extract_fields(extracted):
    return {}


def read_markdown(md_filename):
    md_path = os.path.join(MD_DIR, md_filename)
    if os.path.exists(md_path):
        with open(md_path, "r", encoding="utf-8") as f:
            return f.read()
    return ""


def parse_filename(md_file):
    """Parse 'University_Name__University_Name_Course_Name.md' into university and course."""
    base = md_file.replace(".md", "")
    if "__" in base:
        uni_folder, rest = base.split("__", 1)
        # rest may start with university name again — strip it
        if rest.startswith(uni_folder + "_"):
            course_part = rest[len(uni_folder) + 1:]
        else:
            course_part = rest
    else:
        uni_folder = base
        course_part = base
    uni_name = uni_folder.replace("_", " ")
    course_name = course_part.replace("_", " ")
    return uni_name, course_name


def main():
    if not os.path.exists(MD_DIR):
        print(f"[ERROR] Markdown directory not found: {MD_DIR}")
        return

    md_files = sorted(f for f in os.listdir(MD_DIR) if f.endswith(".md"))
    print(f"Found {len(md_files)} markdown files in {MD_DIR}")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Review ──
    ws1 = wb.active
    ws1.title = "Review"

    headers = [
        "University Name", "course_name",
        "tuition_fees", "application_fee",
        "GRE_required", "GRE_score", "GMAT_required",
        "GPA_min", "work_experience", "LOR_count",
        "TOEFL_min", "IELTS_min", "Duolingo_min", "PTE_min",
        "officialPageLink",
        "confidence_tier",
        "deadline_summer_2026", "deadline_fall_2026", "deadline_winter_2026",
        "deadline_spring_2027", "deadline_summer_2027",
        "deadline_fall_2027", "deadline_winter_2027",
        "deadline_other",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    # ── Sheet 2: Full Markdown ──
    ws2 = wb.create_sheet("Full Markdown")
    md_headers = ["University Name", "course_name", "confidence_tier", "full_markdown"]
    for col, h in enumerate(md_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    print("Processing files...")
    row_num = 2

    for md_file in md_files:
        uni_name, course_name = parse_filename(md_file)

        md_content = read_markdown(md_file)

        # Try to extract a cleaner course name from the first H1 of the markdown
        if md_content:
            h1_match = re.match(r'#\s+([^\n]+)', md_content)
            if h1_match:
                title = h1_match.group(1).strip()
                # Title is usually "University - Course Name" or "University Course Name"
                if " - " in title:
                    parts = title.split(" - ", 1)
                    uni_name = parts[0].strip()
                    course_name = parts[1].strip()
                elif uni_name in title:
                    course_name = title.replace(uni_name, "").strip(" -:")

        # Regex-extract fields from markdown text
        fields = extract_fields_from_md(md_content)
        page_link = fields.get("officialPageLink") or ""

        row_data = [
            uni_name, course_name,
            fields.get("tuition_fees"),
            fields.get("application_fee"),
            fields.get("GRE_required"), fields.get("GRE_score"),
            fields.get("GMAT_required"),
            fields.get("GPA_min"),
            fields.get("work_experience"),
            fields.get("LOR_count"),
            fields.get("TOEFL_min"), fields.get("IELTS_min"),
            fields.get("Duolingo_min"), fields.get("PTE_min"),
            page_link,
            "low",
            fields.get("deadline_summer_2026"),
            fields.get("deadline_fall_2026"),
            fields.get("deadline_winter_2026"),
            fields.get("deadline_spring_2027"),
            fields.get("deadline_summer_2027"),
            fields.get("deadline_fall_2027"),
            fields.get("deadline_winter_2027"),
            fields.get("deadline_other"),
        ]
        for col, val in enumerate(row_data, 1):
            ws1.cell(row=row_num, column=col, value=str(val) if val else "")

        # Sheet 2
        ws2.cell(row=row_num, column=1, value=uni_name)
        ws2.cell(row=row_num, column=2, value=course_name)
        ws2.cell(row=row_num, column=3, value="low")
        ws2.cell(row=row_num, column=4, value=md_content)

        row_num += 1
        if (row_num - 2) % 100 == 0:
            print(f"  Processed {row_num - 2}/{len(md_files)}...")

    # ── Sheet 1 formatting ──
    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    # Wider columns for the deadline columns at the end
    deadline_start_col = headers.index("deadline_summer_2026") + 1
    for col in range(deadline_start_col, len(headers) + 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 35
    ws1.freeze_panes = "A2"  # freeze header row

    # ── Sheet 2 formatting ──
    # Wider columns for readability, and wrap text in markdown column
    ws2.column_dimensions["A"].width = 30   # University Name
    ws2.column_dimensions["B"].width = 40   # course_name
    ws2.column_dimensions["C"].width = 15   # confidence_tier
    ws2.column_dimensions["D"].width = 120  # full_markdown (very wide)

    # Wrap text in markdown column + set row heights
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for row in ws2.iter_rows(min_row=2, max_col=4):
        for cell in row:
            cell.alignment = wrap_align
        # Set a reasonable row height (markdown can be very long)
        ws2.row_dimensions[row[0].row].height = 200

    ws2.freeze_panes = "A2"  # freeze header row

    wb.save(OUTPUT_FILE)
    print(f"\nDone!")
    print(f"  Total rows: {row_num - 2}")
    print(f"  Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
