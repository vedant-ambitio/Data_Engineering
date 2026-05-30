#!/usr/bin/env python3
"""
generate_review_sheet.py — Generate Excel review sheet for moderate & low confidence UG courses
================================================================================================

Sheet 1 (Review): Key fields for quick verification
Sheet 2 (Full Markdown): Complete markdown content per course

Does NOT modify any existing files.
"""

import csv
import json
import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# ── Config ──────────────────────────────────────────────────────────────────

UG_CSV = "ug_programs_data_2026-03-30T18_31_17.92687006+05_30.csv"
EXTRACTED_DIR = "university_data/structured_extraction_ug"
MD_DIRS = {
    "moderate": "classification_results/ug/moderate_confidence",
    "low": "classification_results/ug/low_confidence",
}
OUTPUT_FILE = "ug_review_moderate_low.xlsx"


def load_csv_rows():
    """Load CSV into dict keyed by (university_name, major, specialization)."""
    rows = {}
    with open(UG_CSV, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            uni = row.get("University Name", "").strip()
            major = row.get("course_major_name", "").strip()
            spec = row.get("course_specialization_name", "").strip()
            pid = row.get("program_id", "").strip()
            uid = row.get("university_id", "").strip()
            key = (uni.lower(), major.lower(), spec.lower())
            rows[key] = {
                "program_id": pid,
                "university_id": uid,
                "university_name": uni,
                "course_major_name": major,
                "course_specialization_name": spec,
                "course_level": row.get("course_level", "").strip(),
                "course_degree_name": row.get("course_degree_name", "").strip(),
                "qsRank": row.get("qsRank", "").strip(),
                "officialPageLink": row.get("officialPageLink", "").strip(),
                "officialLinks": row.get("officialLinks", "").strip(),
            }
            # Also store by (uni, spec) for fallback
            key2 = (uni.lower(), spec.lower())
            if key2 not in rows:
                rows[key2] = rows[key]
            # Also by (uni, major)
            key3 = (uni.lower(), major.lower())
            if key3 not in rows:
                rows[key3] = rows[key]
    return rows


def find_csv_match(csv_rows, uni_name, program_name):
    """Find matching CSV row for a course."""
    uni = uni_name.lower().strip()
    prog = program_name.lower().strip()

    # Strip degree prefix
    for prefix in ["bachelor of science in ", "bachelor of arts in ", "bachelor of science ",
                    "bachelor of arts ", "bs in ", "ba in ", "bs ", "ba "]:
        if prog.startswith(prefix):
            prog = prog[len(prefix):]
            break

    # Try (uni, major, spec), (uni, spec), (uni, major)
    for key_uni in [uni, uni.replace(",", "").replace(".", "")]:
        for (u, *rest), row in csv_rows.items():
            if u == key_uni or key_uni in u or u in key_uni:
                combined = " ".join(rest)
                if prog in combined or combined in prog:
                    return row

    # Fallback: word overlap
    for (u, *rest), row in csv_rows.items():
        if uni in u or u in uni:
            combined = " ".join(rest)
            prog_words = set(prog.split()) - {"of", "in", "and", "the", "a", "for", "with"}
            rest_words = set(combined.split()) - {"of", "in", "and", "the", "a", "for", "with"}
            if prog_words and rest_words:
                overlap = len(prog_words & rest_words) / min(len(prog_words), len(rest_words))
                if overlap >= 0.5:
                    return row

    return None


def extract_fields(extracted):
    """Extract key fields from the Gemini extraction JSON."""
    if not extracted:
        return {}

    cost = extracted.get("cost_of_attendance") or {}
    adm = extracted.get("admission_requirements") or {}
    deadlines = extracted.get("deadlines") or []

    # Deadline: earliest fall and spring
    fall_dates = sorted([d["deadline_date"] for d in deadlines
                         if d.get("intake") == "fall" and d.get("deadline_date")])
    spring_dates = sorted([d["deadline_date"] for d in deadlines
                           if d.get("intake") == "spring" and d.get("deadline_date")])

    # TOEFL/IELTS
    toefl_min = None
    ielts_min = None
    for t in (adm.get("english_tests") or []):
        if t.get("test") == "TOEFL" and t.get("min_score"):
            toefl_min = t["min_score"]
        elif t.get("test") == "IELTS" and t.get("min_score"):
            ielts_min = t["min_score"]

    return {
        "tuition_per_year": cost.get("tuition_per_year") or cost.get("tuition_international"),
        "deadline_fall": fall_dates[0] if fall_dates else None,
        "deadline_spring": spring_dates[0] if spring_dates else None,
        "TOEFL_min": toefl_min,
        "IELTS_min": ielts_min,
        "SAT_range": adm.get("sat_score_range"),
        "ACT_range": adm.get("act_score_range"),
        "GPA_min": adm.get("min_gpa"),
    }


def read_markdown(md_filename, tier):
    """Read full markdown content."""
    md_path = os.path.join(MD_DIRS.get(tier, ""), md_filename)
    if os.path.exists(md_path):
        with open(md_path, "r", encoding="utf-8") as f:
            return f.read()
    return ""


def main():
    print("Loading CSV data...")
    csv_rows = load_csv_rows()
    print(f"  {len(csv_rows)} CSV entries loaded")

    # Collect all moderate + low confidence files
    all_files = []
    for tier, md_dir in MD_DIRS.items():
        if not os.path.exists(md_dir):
            print(f"  [WARN] {md_dir} not found")
            continue
        for f in sorted(os.listdir(md_dir)):
            if f.endswith(".md"):
                all_files.append((tier, f))
    print(f"  {len(all_files)} markdown files found (moderate + low)")

    # Create workbook
    wb = openpyxl.Workbook()

    # ── Sheet 1: Review ──
    ws1 = wb.active
    ws1.title = "Review"

    headers = [
        "program_id", "university_id", "University Name",
        "course_major_name", "course_specialization_name",
        "course_level", "course_degree_name", "qsRank",
        "officialPageLink", "officialLinks",
        "deadline_fall", "deadline_spring", "tuition_per_year",
        "TOEFL_min", "IELTS_min", "SAT_range", "ACT_range", "GPA_min",
        "confidence_tier", "markdown_preview",
    ]

    # Style header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    # ── Sheet 2: Full Markdown ──
    ws2 = wb.create_sheet("Full Markdown")
    md_headers = ["program_id", "university_name", "course_name", "confidence_tier", "full_markdown"]
    for col, h in enumerate(md_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    print("Processing files...")
    row_num = 2
    md_row_num = 2
    matched = 0
    unmatched = 0

    for tier, md_file in all_files:
        # Derive university name and program name from filename
        # Format: Uni_Name__Uni_Name_Course.md
        base = md_file.replace(".md", "")
        if "__" in base:
            uni_folder = base.split("__")[0]
        else:
            uni_folder = base

        uni_name_display = uni_folder.replace("_", " ")

        # Load extracted JSON if available
        json_name = md_file.replace(".md", ".json")
        json_path = os.path.join(EXTRACTED_DIR, json_name)
        extracted = {}
        program_name = ""
        if os.path.exists(json_path):
            try:
                with open(json_path, "r", encoding="utf-8") as f:
                    extracted = json.load(f)
                program_name = extracted.get("program_name", "")
                uni_name_display = extracted.get("university_name", uni_name_display)
            except Exception:
                pass

        # Extract key fields
        fields = extract_fields(extracted)

        # Match to CSV
        csv_match = find_csv_match(csv_rows, uni_name_display, program_name or base)

        # Build row
        if csv_match:
            matched += 1
            pid = csv_match["program_id"]
            uid = csv_match["university_id"]
            major = csv_match["course_major_name"]
            spec = csv_match["course_specialization_name"]
            level = csv_match["course_level"]
            degree = csv_match["course_degree_name"]
            qs = csv_match["qsRank"]
            page_link = csv_match["officialPageLink"]
            off_links = csv_match["officialLinks"]
        else:
            unmatched += 1
            pid = ""
            uid = ""
            major = program_name
            spec = ""
            level = extracted.get("degree_type", "")
            degree = extracted.get("degree_type", "")
            qs = ""
            page_link = (extracted.get("important_links") or {}).get("program_page", "")
            off_links = ""

        # Read markdown preview (first 500 chars)
        md_content = read_markdown(md_file, tier)
        preview = md_content[:500].replace("\n", " ") if md_content else ""

        # Write Sheet 1 row
        row_data = [
            pid, uid, uni_name_display,
            major, spec, level, degree, qs,
            page_link, off_links,
            fields.get("deadline_fall"), fields.get("deadline_spring"),
            fields.get("tuition_per_year"),
            fields.get("TOEFL_min"), fields.get("IELTS_min"),
            fields.get("SAT_range"), fields.get("ACT_range"),
            fields.get("GPA_min"),
            tier, preview,
        ]
        for col, val in enumerate(row_data, 1):
            ws1.cell(row=row_num, column=col, value=str(val) if val else "")
        row_num += 1

        # Write Sheet 2 row
        ws2.cell(row=md_row_num, column=1, value=str(pid))
        ws2.cell(row=md_row_num, column=2, value=uni_name_display)
        ws2.cell(row=md_row_num, column=3, value=program_name or major)
        ws2.cell(row=md_row_num, column=4, value=tier)
        ws2.cell(row=md_row_num, column=5, value=md_content)
        md_row_num += 1

        if (row_num - 2) % 500 == 0:
            print(f"  Processed {row_num - 2} / {len(all_files)}...")

    # Auto-width for Sheet 1 (approximate)
    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    # Narrower for preview
    ws1.column_dimensions[openpyxl.utils.get_column_letter(len(headers))].width = 60

    # Save
    wb.save(OUTPUT_FILE)
    print(f"\nDone!")
    print(f"  Total rows: {row_num - 2}")
    print(f"  Matched to CSV: {matched}")
    print(f"  Unmatched: {unmatched}")
    print(f"  Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
