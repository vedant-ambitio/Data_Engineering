"""
Patch source .md files (masters_data/masters/* and phd_data/phd/*) with
manually verified corrections from courses_patch/*.xlsx.

For each Excel row with Notes filled:
  1. Locate the source .md file via H1 title match.
  2. If Excel 'full_markdown' column differs from disk -> overwrite with Excel content.
  3. Otherwise build a "Reviewer-Verified" insertion in the Tuition & Fees
     and Admission Requirements sections containing the verified values
     (Tuition fee Verified, Application fee, Stipend verified, link cols).

Writes patched_files.json (list of {kind, source_path, reasons}) for the
re-classify step.
"""

import json
import re
from pathlib import Path

import openpyxl

URL_RE = re.compile(r'^https?://', re.IGNORECASE)


def _is_url(value: str) -> bool:
    return bool(value and URL_RE.match(str(value).strip()))

BASE = Path(r"c:\Users\HP\OneDrive\Desktop\course_data")
PATCH_DIR = BASE / "courses_patch"
MASTERS_XLSX = PATCH_DIR / "masters_review_low_v7.xlsx"
PHD_XLSX = PATCH_DIR / "phd_review_low_v2.xlsx"
MASTERS_MD_DIR = BASE / "masters_data" / "masters"
PHD_MD_DIR = BASE / "phd_data" / "phd"
LOG_PATH = BASE / "courses_patch" / "patched_files.json"


# ─────────────────────────────────────────────
# Build H1 -> path index
# ─────────────────────────────────────────────

def build_h1_index(md_root: Path):
    """Return {h1_title: Path}.  H1 is the first line starting with '# '."""
    idx = {}
    for college_dir in md_root.iterdir():
        if not college_dir.is_dir():
            continue
        for md in college_dir.glob("*.md"):
            try:
                with md.open("r", encoding="utf-8", errors="replace") as f:
                    head = f.read(500)
            except Exception:
                continue
            m = re.match(r"#\s+([^\n]+)", head)
            if not m:
                continue
            h1 = m.group(1).strip()
            # Keep first occurrence; collisions are rare and not worth solving
            idx.setdefault(h1, md)
    return idx


# ─────────────────────────────────────────────
# Excel row helpers
# ─────────────────────────────────────────────

def get_h1_from_md(md_text: str) -> str | None:
    if not md_text:
        return None
    m = re.match(r"#\s+([^\n]+)", md_text)
    return m.group(1).strip() if m else None


def build_correction_block_masters(row, link_uni_official: str | None) -> str:
    """
    Build the Reviewer-Verified Correction block to append to ## Tuition & Fees
    from a Masters Review row.

      0 Notes
      1 University Name
      2 course_name
      3 tuition_fees (orig)
      4 Tuition fee Links (URL)
      5 Tuition fee Verified
      6 application_fee (orig)
      7 Application fee (corrected)
      18 officialPageLink
    """
    notes = (row[0] or "").strip()
    tuition_verified = (str(row[5]).strip() if len(row) > 5 and row[5] is not None else "")
    tuition_link = (str(row[4]).strip() if len(row) > 4 and row[4] is not None else "")
    app_fee_raw = (str(row[7]).strip() if len(row) > 7 and row[7] is not None else "")

    # Distinguish URL vs corrected value for the app-fee column.
    app_fee_value = "" if _is_url(app_fee_raw) else app_fee_raw
    app_fee_url = app_fee_raw if _is_url(app_fee_raw) else ""

    if not tuition_verified and not app_fee_value and not app_fee_url:
        return ""

    lines = ["", "### Reviewer-Verified Correction", f"_Reviewer note: {notes}_", ""]
    if tuition_verified:
        lines.append(f"*   **Verified Tuition Fee:** {tuition_verified}")
    if app_fee_value:
        lines.append(f"*   **Verified Application Fee:** {app_fee_value}")
    if tuition_link:
        lines.append(f"*   **Tuition Source URL:** {tuition_link}")
    if app_fee_url:
        lines.append(f"*   **Application Fee Source URL:** {app_fee_url}")
    if link_uni_official and not (tuition_link or app_fee_url):
        lines.append(f"*   **Source URL:** {link_uni_official}")
    lines.append("")
    return "\n".join(lines)


def build_correction_block_phd(row) -> str:
    """
    PhD Review row layout:
      0 Notes
      1 Uni  2 Course
      3 tuition_fees (orig)
      4 (nameless, often Tuition fee Links)
      5 Tuition fee verified
      6 application_fee (orig)
      7 application fee (corrected / URL)
      8 stipend (orig)
      9 stipend verified (value or URL)
     22 officialPageLink
    """
    notes = (row[0] or "").strip()
    tuition_link = (str(row[4]).strip() if len(row) > 4 and row[4] is not None else "")
    tuition_verified = (str(row[5]).strip() if len(row) > 5 and row[5] is not None else "")
    app_fee_raw = (str(row[7]).strip() if len(row) > 7 and row[7] is not None else "")
    stipend_raw = (str(row[9]).strip() if len(row) > 9 and row[9] is not None else "")
    official_url = (str(row[22]).strip() if len(row) > 22 and row[22] is not None else "")

    app_fee_value = "" if _is_url(app_fee_raw) else app_fee_raw
    app_fee_url = app_fee_raw if _is_url(app_fee_raw) else ""
    stipend_value = "" if _is_url(stipend_raw) else stipend_raw
    stipend_url = stipend_raw if _is_url(stipend_raw) else ""

    has_data = any([tuition_verified, app_fee_value, app_fee_url, stipend_value, stipend_url])
    if not has_data:
        return ""

    lines = ["", "### Reviewer-Verified Correction", f"_Reviewer note: {notes}_", ""]
    if tuition_verified:
        lines.append(f"*   **Verified Tuition Fee:** {tuition_verified}")
    if app_fee_value:
        lines.append(f"*   **Verified Application Fee:** {app_fee_value}")
    if stipend_value:
        lines.append(f"*   **Verified Stipend / Funding:** {stipend_value}")
    if tuition_link:
        lines.append(f"*   **Tuition Source URL:** {tuition_link}")
    if app_fee_url:
        lines.append(f"*   **Application Fee Source URL:** {app_fee_url}")
    if stipend_url:
        lines.append(f"*   **Stipend Source URL:** {stipend_url}")
    if official_url and not (tuition_link or app_fee_url or stipend_url):
        lines.append(f"*   **Official Page:** {official_url}")
    lines.append("")
    return "\n".join(lines)


# ─────────────────────────────────────────────
# Markdown surgery
# ─────────────────────────────────────────────

def append_to_section(md_text: str, section_heading: str, block: str) -> str:
    """
    Append `block` at the end of the named `## section_heading` section
    (before the next `## ` heading or EOF), but AFTER any closing </citation>.
    Returns the modified text.  No-op if section not found.
    """
    if not block:
        return md_text
    pattern = rf'(^## {re.escape(section_heading)}\s*\n.*?)(?=\n## [A-Z]|\Z)'
    m = re.search(pattern, md_text, flags=re.DOTALL | re.MULTILINE)
    if not m:
        return md_text
    body = m.group(1).rstrip()
    # Strip a trailing duplicate "Reviewer-Verified Correction" block (idempotent re-runs)
    body = re.sub(
        r'\n+### Reviewer-Verified Correction.*?(?=\n## |\Z)',
        '',
        body,
        flags=re.DOTALL,
    ).rstrip()
    new_body = body + "\n" + block
    return md_text[:m.start()] + new_body + md_text[m.end():]


_BLOCK_STRIP_RE = re.compile(
    r'\n+### Reviewer-Verified Correction.*?(?=\n## |\Z)',
    flags=re.DOTALL,
)


def _strip_block(text: str) -> str:
    """Remove any Reviewer-Verified Correction block previously inserted by this script."""
    return _BLOCK_STRIP_RE.sub('', text or '')


def patch_one(md_path: Path, excel_full_md: str, correction_block: str) -> tuple[bool, str]:
    """
    Patch the source md file.

    Returns (changed, reason).
    """
    try:
        disk = md_path.read_text(encoding="utf-8", errors="replace")
    except FileNotFoundError:
        return False, "source_missing"

    # Case 1 — reviewer edited the full markdown in Excel: take it verbatim.
    # Compare AFTER stripping any prior Reviewer-Verified block from disk; otherwise our
    # own previous insertion would be misread as "reviewer edited the Excel".
    disk_stripped = _strip_block(disk).strip()
    excel_stripped = _strip_block(excel_full_md or "").strip()
    if excel_stripped and excel_stripped != disk_stripped:
        if get_h1_from_md(excel_full_md) and "## Tuition" in excel_full_md:
            md_path.write_text(excel_full_md, encoding="utf-8")
            return True, "replaced_with_excel_full_markdown"

    # Case 2 — append correction block to Tuition & Fees.
    if not correction_block:
        return False, "no_correction_data"

    new_text = append_to_section(disk, "Tuition & Fees", correction_block)
    if new_text == disk:
        # Section heading wasn't found — append a synthetic Tuition section at EOF
        new_text = disk.rstrip() + "\n\n## Tuition & Fees\n" + correction_block + "\n"
    if new_text == disk:
        return False, "no_change"
    md_path.write_text(new_text, encoding="utf-8")
    return True, "appended_correction_block"


# ─────────────────────────────────────────────
# Main per-kind driver
# ─────────────────────────────────────────────

def patch_kind(kind: str, xlsx_path: Path, md_root: Path) -> list[dict]:
    print(f"\n=== Patching {kind} from {xlsx_path.name} ===")
    print(f"Building H1 index from {md_root}...", flush=True)
    h1_idx = build_h1_index(md_root)
    print(f"  Indexed {len(h1_idx)} files")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    review = list(wb["Review"].iter_rows(min_row=2, values_only=True))
    fullmd = list(wb["Full Markdown"].iter_rows(min_row=2, values_only=True))
    assert len(review) == len(fullmd), "Sheet row counts disagree"

    patched = []
    no_notes = 0
    no_match = 0
    no_change = 0

    for r, m in zip(review, fullmd):
        notes = r[0] if r and len(r) > 0 else None
        if not notes:
            no_notes += 1
            continue

        excel_md = m[3] if m and len(m) > 3 else ""
        h1 = get_h1_from_md(excel_md or "")
        if not h1:
            # fallback: build from uni + course
            uni = (m[0] if m and len(m) > 0 else None) or r[1]
            course = (m[1] if m and len(m) > 1 else None) or r[2]
            h1 = f"{uni} - {course}" if uni and course else None
        if not h1 or h1 not in h1_idx:
            no_match += 1
            continue

        src_path = h1_idx[h1]
        if kind == "masters":
            link_uni = str(r[18]).strip() if len(r) > 18 and r[18] else None
            block = build_correction_block_masters(r, link_uni)
        else:
            block = build_correction_block_phd(r)

        changed, reason = patch_one(src_path, excel_md or "", block)
        if changed:
            patched.append({
                "kind": kind,
                "source_path": str(src_path),
                "h1": h1,
                "notes": notes,
                "reason": reason,
            })
        else:
            no_change += 1

    print(f"  Patched: {len(patched)}")
    print(f"  Skipped (no notes):   {no_notes}")
    print(f"  Skipped (no h1 match): {no_match}")
    print(f"  Skipped (no change):   {no_change}")
    return patched


def main():
    log = []
    log += patch_kind("masters", MASTERS_XLSX, MASTERS_MD_DIR)
    log += patch_kind("phd", PHD_XLSX, PHD_MD_DIR)

    LOG_PATH.write_text(json.dumps(log, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\nWrote patch log: {LOG_PATH}  ({len(log)} entries)")


if __name__ == "__main__":
    main()
