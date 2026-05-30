"""Break down WHY entries are NOT_FOUND in the v3 report."""
import json
import os
from collections import Counter, defaultdict

from openpyxl import load_workbook

BASE = r"c:\Users\HP\OneDrive\Desktop\course_data\Professors_info"
WB = os.path.join(BASE, "coverage_grounding_v3.xlsx")

# Load drill-down rows
wb = load_workbook(WB)
ws = wb["Missing Universities"]
rows = list(ws.iter_rows(values_only=True))[1:]

# Reason histogram
reasons = Counter()
for r in rows:
    dept, uni, total_ch, off_ch, off_work, reason, url = r
    # Bucket the reasons
    if reason == "no grounding chunks":
        reasons["no grounding chunks"] += 1
    elif reason == "chunks present but none on official domain":
        reasons["chunks but none on official domain"] += 1
    elif "official chunk(s) but none probed working" in (reason or ""):
        reasons["official chunks but probe failed"] += 1
    else:
        reasons["other"] += 1

total = len(rows)
print(f"Total NOT_FOUND: {total}\n")
print("Breakdown by reason:")
for r, n in reasons.most_common():
    print(f"  {n:5d}  ({n/total*100:5.1f}%)  {r}")

# Cross-reference with original confidence
print("\nCross-referencing with original confidence (from output/universities/)...")
output_dir = os.path.join(BASE, "output", "universities")

# Load output data
orig_conf = {}  # (uni, dept) -> confidence
for fn in os.listdir(output_dir):
    if not fn.endswith(".json"):
        continue
    with open(os.path.join(output_dir, fn), "r", encoding="utf-8") as f:
        data = json.load(f)
    uni = data.get("university", "")
    for dept, entry in data.get("departments", {}).items():
        orig_conf[(uni, dept)] = {
            "confidence": entry.get("confidence"),
            "pre_filtered": entry.get("pre_filtered", False),
            "verification_status": entry.get("verification_status"),
        }

# Cross-tab: NOT_FOUND reason × original confidence
crosstab = defaultdict(Counter)
for r in rows:
    dept, uni, total_ch, off_ch, off_work, reason, url = r
    info = orig_conf.get((uni, dept), {})
    conf = info.get("confidence") or "missing"
    pf = info.get("pre_filtered")

    if reason == "no grounding chunks":
        bucket = "no chunks"
    elif reason == "chunks present but none on official domain":
        bucket = "non-official chunks"
    elif "official chunk(s)" in (reason or ""):
        bucket = "official, no probe"
    else:
        bucket = "other"

    key = f"{conf}{' (pre-filtered)' if pf else ''}"
    crosstab[bucket][key] += 1

print()
for bucket, counts in crosstab.items():
    print(f"\n{bucket} ({sum(counts.values())} records) -- by original confidence:")
    for conf, n in counts.most_common():
        print(f"  {n:5d}  {conf}")

# Also: how many of the "no grounding chunks" entries were pre_filtered?
print("\n=== Of the 'no grounding chunks' bucket ===")
no_chunks_rows = [r for r in rows if r[5] == "no grounding chunks"]
pre_filtered = sum(1 for r in no_chunks_rows
                   if orig_conf.get((r[1], r[0]), {}).get("pre_filtered"))
print(f"  pre_filtered=True: {pre_filtered}")
print(f"  pre_filtered=False: {len(no_chunks_rows) - pre_filtered}")
