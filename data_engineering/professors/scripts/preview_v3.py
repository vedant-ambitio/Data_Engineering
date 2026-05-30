from openpyxl import load_workbook
wb = load_workbook(r"c:\Users\HP\OneDrive\Desktop\course_data\Professors_info\coverage_grounding_v3.xlsx")
ws = wb["Coverage by Department"]
print(f"{'Department':<38} {'Tot':>5} {'FND':>5} {'NF':>5} {'Cov%':>6}")
print("-" * 65)
for r in list(ws.iter_rows(values_only=True))[1:]:
    dept, total, found, nf, cov = r
    print(f"{dept:<38} {total:>5} {found:>5} {nf:>5} {cov:>6}")
