"""Coverage analysis based on GROUNDING CHUNKS (Tier A only).

Definition of FOUND for a (university, department):
  - At least one entry in grounding_metadata.groundingChunks whose `domain`
    matches the university's official domain (or a subdomain of it), AND
  - url_verification.verification_status is "ok" or "ok_via_alternate"

Otherwise NOT_FOUND. Original confidence labels (high/low/not_found) are ignored.
"""
import csv
import json
import os
from collections import defaultdict
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = r"c:\Users\HP\OneDrive\Desktop\course_data\Professors_info"
OUTPUT_UNIS = os.path.join(BASE, "output", "universities")
GROUNDING_DIR = os.path.join(BASE, "grounding")
TIER_A_CSV = os.path.join(BASE, "config", "universities_tier_A.csv")
OUT_XLSX = os.path.join(BASE, "coverage_grounding_tierA.xlsx")


# Load Tier A: name -> official_website
tier_a = {}
with open(TIER_A_CSV, "r", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        name = row["university_name"].strip()
        site = row.get("official_website", "").strip()
        if name:
            tier_a[name] = site


def to_domain(url_or_host):
    if not url_or_host:
        return ""
    s = url_or_host.strip().lower()
    if "://" in s:
        try:
            s = urlparse(s).hostname or ""
        except Exception:
            return ""
    if s.startswith("www."):
        s = s[4:]
    return s.strip(".")


def is_subdomain_match(chunk_domain, official_domain):
    cd = to_domain(chunk_domain)
    od = to_domain(official_domain)
    if not cd or not od:
        return False
    return cd == od or cd.endswith("." + od)


def analyze_grounding(grounding_path, official_domain):
    """Return dict with status/details for one (uni, dept)."""
    if not grounding_path or not os.path.exists(grounding_path):
        return {"status": "NOT_FOUND", "has_official_chunk": False,
                "verified": False, "working_url": "", "reason": "no grounding file"}
    try:
        with open(grounding_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        return {"status": "NOT_FOUND", "has_official_chunk": False,
                "verified": False, "working_url": "", "reason": f"parse error: {e}"}

    od = official_domain or data.get("university_domain", "")
    chunks = (data.get("grounding_metadata") or {}).get("groundingChunks") or []
    has_official_chunk = False
    matched_chunk_domain = ""
    for c in chunks:
        web = (c or {}).get("web") or {}
        chunk_domain = web.get("domain") or to_domain(web.get("uri", ""))
        if is_subdomain_match(chunk_domain, od):
            has_official_chunk = True
            matched_chunk_domain = chunk_domain
            break

    uv = data.get("url_verification") or {}
    vstatus = uv.get("verification_status")
    verified = vstatus in ("ok", "ok_via_alternate")
    working_url = uv.get("verified_url") if verified else ""

    if has_official_chunk and verified:
        return {"status": "FOUND", "has_official_chunk": True, "verified": True,
                "working_url": working_url or "", "reason": ""}

    # Build a useful reason for the drill-down
    if not chunks:
        reason = "no grounding chunks"
    elif not has_official_chunk:
        reason = "chunks present but none on official domain"
    elif not verified:
        reason = f"chunks ok, URL not verified (status={vstatus})"
    else:
        reason = "unknown"
    return {"status": "NOT_FOUND", "has_official_chunk": has_official_chunk,
            "verified": verified, "working_url": working_url or "", "reason": reason}


# Walk every Tier A university JSON in output/universities/
records = []
unis_processed = set()
unis_skipped_nontier = []

for fn in sorted(os.listdir(OUTPUT_UNIS)):
    if not fn.endswith(".json"):
        continue
    fp = os.path.join(OUTPUT_UNIS, fn)
    with open(fp, "r", encoding="utf-8") as f:
        data = json.load(f)
    uni = data.get("university", "")
    if uni not in tier_a:
        unis_skipped_nontier.append(uni)
        continue
    unis_processed.add(uni)
    official_domain = tier_a[uni] or data.get("university_domain", "")
    for dept, entry in data.get("departments", {}).items():
        gfile = entry.get("grounding_file")
        gpath = os.path.join(GROUNDING_DIR, gfile) if gfile else None
        result = analyze_grounding(gpath, official_domain)
        records.append({
            "university": uni,
            "department": dept,
            "status": result["status"],
            "has_official_chunk": result["has_official_chunk"],
            "verified": result["verified"],
            "working_url": result["working_url"],
            "reason": result["reason"],
        })

print(f"Tier A in CSV:               {len(tier_a)}")
print(f"Tier A unis processed:       {len(unis_processed)}")
print(f"Output unis NOT in Tier A:   {len(unis_skipped_nontier)}")
print(f"Total (uni, dept) records:   {len(records)}")


# Aggregate per department
dept_stats = defaultdict(lambda: {"total": 0, "FOUND": 0, "NOT_FOUND": 0})
for r in records:
    s = dept_stats[r["department"]]
    s["total"] += 1
    s[r["status"]] += 1


a_rows = []
for dept, s in dept_stats.items():
    cov = (s["FOUND"] / s["total"] * 100.0) if s["total"] else 0.0
    a_rows.append({
        "Department": dept,
        "Tier A universities": s["total"],
        "FOUND (official+working)": s["FOUND"],
        "NOT FOUND": s["NOT_FOUND"],
        "Coverage %": round(cov, 1),
    })
a_rows.sort(key=lambda r: r["Coverage %"])

# Drill-down: every NOT FOUND record, ordered by worst dept first
dept_order = {row["Department"]: i for i, row in enumerate(a_rows)}
c_rows = [r for r in records if r["status"] == "NOT_FOUND"]
c_rows.sort(key=lambda r: (dept_order.get(r["department"], 999), r["university"]))


# Build Excel
wb = Workbook()

# Summary sheet
ws_sum = wb.active
ws_sum.title = "Summary"
ws_sum.append(["Coverage by Grounding Chunks — Tier A only"])
ws_sum["A1"].font = Font(bold=True, size=14)

total = len(records)
total_found = sum(1 for r in records if r["status"] == "FOUND")
total_nf = total - total_found

ws_sum.append([])
ws_sum.append(["Tier A universities processed", len(unis_processed)])
ws_sum.append(["Total (uni, dept) records", total])
ws_sum.append(["FOUND (official chunk + working URL)", total_found, f"{total_found/total*100:.1f}%"])
ws_sum.append(["NOT FOUND", total_nf, f"{total_nf/total*100:.1f}%"])
ws_sum.append([])
ws_sum.append(["FOUND criterion:"])
ws_sum.append(["", "1. groundingChunks contains at least one chunk on the official domain"])
ws_sum.append(["", "2. url_verification.verification_status is 'ok' or 'ok_via_alternate'"])
ws_sum.append([])
ws_sum.append(["Sheets:"])
ws_sum.append(["", "1. Coverage by Department  — sorted worst → best"])
ws_sum.append(["", "2. Missing Universities    — every NOT FOUND, filterable by Department"])
ws_sum.column_dimensions["A"].width = 40
ws_sum.column_dimensions["B"].width = 18
ws_sum.column_dimensions["C"].width = 12


# Sheet: Coverage by Department
ws1 = wb.create_sheet("Coverage by Department")
headers_a = ["Department", "Tier A universities", "FOUND (official+working)",
             "NOT FOUND", "Coverage %"]
ws1.append(headers_a)
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="305496")
for i in range(1, len(headers_a) + 1):
    c = ws1.cell(row=1, column=i)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[1].height = 30

red = PatternFill("solid", fgColor="F8CBAD")
amber = PatternFill("solid", fgColor="FFE699")
green = PatternFill("solid", fgColor="C6E0B4")

for r in a_rows:
    ws1.append([r[h] for h in headers_a])
    last = ws1.max_row
    cov = r["Coverage %"]
    fill = red if cov < 50 else amber if cov < 80 else green
    ws1.cell(row=last, column=5).fill = fill

for i, w in enumerate([40, 16, 18, 14, 12], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = ws1.dimensions


# Sheet: Missing Universities (drill-down)
ws2 = wb.create_sheet("Missing Universities")
headers_c = ["Department", "University", "Has official-domain chunk?",
             "URL verified?", "Reason", "Best URL we have"]
ws2.append(headers_c)
for i in range(1, len(headers_c) + 1):
    c = ws2.cell(row=1, column=i)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")

for r in c_rows:
    ws2.append([
        r["department"],
        r["university"],
        "yes" if r["has_official_chunk"] else "no",
        "yes" if r["verified"] else "no",
        r["reason"],
        r["working_url"],
    ])

for i, w in enumerate([35, 45, 18, 14, 40, 70], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions


wb.save(OUT_XLSX)
print(f"\nWrote {OUT_XLSX}")
print(f"  - {len(a_rows)} departments")
print(f"  - {len(c_rows)} NOT FOUND rows in drill-down")
print(f"\nFOUND total: {total_found} / {total} ({total_found/total*100:.1f}%)")
