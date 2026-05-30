"""Build Excel coverage report:
- Sheet 1: Coverage by Department (Option A)
- Sheet 2: Missing Universities drill-down (Option C)
HIGH = original high + recovered (combined into one column).
"""
import csv
import json
import os
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = r"c:\Users\HP\OneDrive\Desktop\course_data\Professors_info"
OUTPUT_DIR = os.path.join(BASE, "output", "universities")
RECOVERY_DIR = os.path.join(BASE, "output_recovery", "universities")
CONFIG_DIR = os.path.join(BASE, "config")

OUT_XLSX = os.path.join(BASE, "coverage_report.xlsx")


def load_tier(path):
    s = set()
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get("university_name", "").strip()
            if name:
                s.add(name)
    return s


tier_a = load_tier(os.path.join(CONFIG_DIR, "universities_tier_A.csv"))
tier_b = load_tier(os.path.join(CONFIG_DIR, "universities_tier_B.csv"))
tier_c = load_tier(os.path.join(CONFIG_DIR, "universities_tier_C.csv"))


def tier_of(uni):
    if uni in tier_a:
        return "A"
    if uni in tier_b:
        return "B"
    if uni in tier_c:
        return "C"
    return "?"


# Load original output
output_data = {}  # uni -> {dept -> entry}
for fn in sorted(os.listdir(OUTPUT_DIR)):
    if not fn.endswith(".json"):
        continue
    with open(os.path.join(OUTPUT_DIR, fn), "r", encoding="utf-8") as f:
        data = json.load(f)
    output_data[data.get("university", fn.replace(".json", ""))] = data.get("departments", {})

# Load recovery
recovery_data = {}
for fn in sorted(os.listdir(RECOVERY_DIR)):
    if not fn.endswith(".json"):
        continue
    with open(os.path.join(RECOVERY_DIR, fn), "r", encoding="utf-8") as f:
        data = json.load(f)
    recovery_data[data.get("university", fn.replace(".json", ""))] = data.get("departments", {})


# Compute merged status per (uni, dept)
# status in {"HIGH", "LOW", "NOT_FOUND", "OTHER"}
records = []  # list of (uni, dept, status, original_conf, url, tier)
for uni, depts in output_data.items():
    rec_depts = recovery_data.get(uni, {})
    for dept, entry in depts.items():
        conf = entry.get("confidence")
        url = entry.get("faculty_page_url")
        rec = rec_depts.get(dept)
        recovered_url = None
        if rec and rec.get("recovery_outcome") == "recovered" and rec.get("faculty_page_url"):
            recovered_url = rec["faculty_page_url"]

        if conf == "high":
            status = "HIGH"
            final_url = url
        elif recovered_url:
            status = "HIGH"  # treat recovered as high per user request
            final_url = recovered_url
        elif conf == "low":
            status = "LOW"
            final_url = url
        elif conf == "medium":
            status = "LOW"  # bucket medium with low (small count: 67)
            final_url = url
        elif conf == "not_found":
            status = "NOT_FOUND"
            final_url = None
        else:
            status = "OTHER"
            final_url = url

        records.append({
            "university": uni,
            "department": dept,
            "status": status,
            "original_conf": conf,
            "url": final_url,
            "tier": tier_of(uni),
        })


# Aggregate per department
dept_stats = defaultdict(lambda: {
    "total": 0, "HIGH": 0, "LOW": 0, "NOT_FOUND": 0, "OTHER": 0,
    "tierA_total": 0, "tierA_high": 0,
    "tierBC_total": 0, "tierBC_high": 0,
})

for r in records:
    d = dept_stats[r["department"]]
    d["total"] += 1
    d[r["status"]] += 1
    if r["tier"] == "A":
        d["tierA_total"] += 1
        if r["status"] == "HIGH":
            d["tierA_high"] += 1
    else:
        d["tierBC_total"] += 1
        if r["status"] == "HIGH":
            d["tierBC_high"] += 1


# Build Option A rows
a_rows = []
for dept, s in dept_stats.items():
    offering = s["total"] - s["NOT_FOUND"]  # universities where dept exists
    coverage_pct = (s["HIGH"] / offering * 100.0) if offering else 0.0
    tierA_offering = s["tierA_total"] - 0  # not_found split not tracked per tier; approximate
    # Use total tier A as denominator for coverage – include not_found separately so we don't penalise
    # Actually for fair comparison let's keep denominator = tier total (offering not separable per tier without re-walk)
    tierA_cov = (s["tierA_high"] / s["tierA_total"] * 100.0) if s["tierA_total"] else 0.0
    tierBC_cov = (s["tierBC_high"] / s["tierBC_total"] * 100.0) if s["tierBC_total"] else 0.0
    a_rows.append({
        "Department": dept,
        "Universities offering it": offering,
        "HIGH (working)": s["HIGH"],
        "LOW (missing)": s["LOW"] + s["OTHER"],
        "NOT_FOUND": s["NOT_FOUND"],
        "Coverage %": round(coverage_pct, 1),
        "Tier A coverage %": round(tierA_cov, 1),
        "Tier B/C coverage %": round(tierBC_cov, 1),
        "Total entries": s["total"],
    })

# Sort ascending by coverage %
a_rows.sort(key=lambda r: r["Coverage %"])


# Build Option C rows (drill-down: long format, every (dept, uni) with status LOW/OTHER)
c_rows = []
for r in records:
    if r["status"] in ("LOW", "OTHER"):
        c_rows.append({
            "Department": r["department"],
            "University": r["university"],
            "Tier": r["tier"],
            "Status": r["status"],
            "Original confidence": r["original_conf"],
            "Best URL we have": r["url"] or "",
        })

# Sort drill-down: by department (worst-coverage first), then university
dept_order = {row["Department"]: i for i, row in enumerate(a_rows)}
c_rows.sort(key=lambda r: (dept_order.get(r["Department"], 999), r["University"]))


# Write Excel
wb = Workbook()

# Sheet 1: Option A
ws1 = wb.active
ws1.title = "Coverage by Department"
headers_a = ["Department", "Universities offering it", "HIGH (working)",
             "LOW (missing)", "NOT_FOUND", "Coverage %",
             "Tier A coverage %", "Tier B/C coverage %", "Total entries"]
ws1.append(headers_a)

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="305496")
for col_idx, _ in enumerate(headers_a, 1):
    c = ws1.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[1].height = 30

red_fill = PatternFill("solid", fgColor="F8CBAD")     # <50%
amber_fill = PatternFill("solid", fgColor="FFE699")   # 50-80%
green_fill = PatternFill("solid", fgColor="C6E0B4")   # >=80%

for r in a_rows:
    row_vals = [r[h] for h in headers_a]
    ws1.append(row_vals)
    last = ws1.max_row
    cov = r["Coverage %"]
    if cov < 50:
        fill = red_fill
    elif cov < 80:
        fill = amber_fill
    else:
        fill = green_fill
    ws1.cell(row=last, column=6).fill = fill  # Coverage %

# Column widths
widths_a = [40, 12, 14, 14, 12, 12, 14, 16, 12]
for i, w in enumerate(widths_a, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = ws1.dimensions

# Sheet 2: Option C drill-down
ws2 = wb.create_sheet("Missing Universities")
headers_c = ["Department", "University", "Tier", "Status", "Original confidence", "Best URL we have"]
ws2.append(headers_c)
for col_idx, _ in enumerate(headers_c, 1):
    c = ws2.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")

for r in c_rows:
    ws2.append([r[h] for h in headers_c])

widths_c = [35, 45, 6, 12, 18, 80]
for i, w in enumerate(widths_c, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions


# Sheet 3: Summary
ws3 = wb.create_sheet("Summary", 0)  # insert at front
ws3.append(["Coverage Report Summary"])
ws3["A1"].font = Font(bold=True, size=14)

total_entries = sum(s["total"] for s in dept_stats.values())
total_high = sum(s["HIGH"] for s in dept_stats.values())
total_low = sum(s["LOW"] + s["OTHER"] for s in dept_stats.values())
total_nf = sum(s["NOT_FOUND"] for s in dept_stats.values())
total_offering = total_entries - total_nf

ws3.append([])
ws3.append(["Total dept entries", total_entries])
ws3.append(["HIGH (working, incl. recovered)", total_high, f"{total_high/total_entries*100:.1f}%"])
ws3.append(["LOW (missing)", total_low, f"{total_low/total_entries*100:.1f}%"])
ws3.append(["NOT_FOUND (dept doesn't exist)", total_nf, f"{total_nf/total_entries*100:.1f}%"])
ws3.append([])
ws3.append(["Coverage % (HIGH / dept-offerings)", f"{total_high/total_offering*100:.1f}%"])
ws3.append([])
ws3.append(["Universities covered", len(output_data)])
ws3.append(["Departments seen", len(dept_stats)])
ws3.append([])
ws3.append(["Sheets:"])
ws3.append(["", "1. Coverage by Department  — sorted worst → best"])
ws3.append(["", "2. Missing Universities    — drill-down per dept/uni"])
ws3.column_dimensions["A"].width = 38
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 12

wb.save(OUT_XLSX)
print(f"Wrote {OUT_XLSX}")
print(f"  - {len(a_rows)} departments in coverage table")
print(f"  - {len(c_rows)} (dept, uni) rows in drill-down")
