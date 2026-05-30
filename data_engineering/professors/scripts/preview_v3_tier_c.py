from openpyxl import load_workbook
wb = load_workbook(r"c:\Users\HP\OneDrive\Desktop\course_data\Professors_info\coverage_grounding_v3_tier_c.xlsx")
ws = wb["Coverage by Department"]
print(f"{'Department':<38} {'Tot':>4} {'FND':>4} {'NF':>4} {'Cov%':>6} {'Absent':>7}")
print("-" * 72)
for r in list(ws.iter_rows(values_only=True))[1:]:
    dept, tot, fnd, nf, cov, absent = r
    print(f"{dept:<38} {tot:>4} {fnd:>4} {nf:>4} {cov:>6} {absent:>7}")
