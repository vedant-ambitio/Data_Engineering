"""Stage 3 (Tier C): classify each (uni, dept) using unwrap + probe caches and write Excel.

Source:
  grounding_tier_c/             (690 grounding files)
  logs/state_tier_c.jsonl       (per-pair confidence info)
  unwrap_cache_tier_c.json
  probe_cache_tier_c.json
  config/universities_tier_C.csv

Output: coverage_grounding_v3_tier_c.xlsx
"""
import csv
import json
import os
from collections import defaultdict
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = r"c:\Users\HP\OneDrive\Desktop\course_data\Professors_info"
GROUNDING = os.path.join(BASE, "grounding_tier_c")
STATE_LOG = os.path.join(BASE, "logs", "state_tier_c.jsonl")
TIER_CSV = os.path.join(BASE, "config", "universities_tier_C.csv")
UNWRAP_CACHE = os.path.join(BASE, "unwrap_cache_tier_c.json")
PROBE_CACHE = os.path.join(BASE, "probe_cache_tier_c.json")
OUT_XLSX = os.path.join(BASE, "coverage_grounding_v3_tier_c.xlsx")


def to_domain(url_or_host):
    if not url_or_host:
        return ""
    s = str(url_or_host).strip().lower()
    if "://" in s:
        try:
            s = urlparse(s).hostname or ""
        except Exception:
            return ""
    if s.startswith("www."):
        s = s[4:]
    return s.strip(".")


def is_subdomain_match(chunk_domain, official_domain):
    cd = to_domain(chunk_domain)
    od = to_domain(official_domain)
    if not cd or not od:
        return False
    return cd == od or cd.endswith("." + od)


# Tier C universities CSV
tier = {}
with open(TIER_CSV, "r", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        name = row["university_name"].strip()
        site = row.get("official_website", "").strip()
        if name:
            tier[name] = site

print(f"Tier C universities in CSV: {len(tier)}")

# Caches
print("Loading caches...")
with open(UNWRAP_CACHE, "r", encoding="utf-8") as f:
    unwrap = json.load(f)
with open(PROBE_CACHE, "r", encoding="utf-8") as f:
    probe = json.load(f)
print(f"  unwrap: {len(unwrap)}  probe: {len(probe)}")

# Original confidences from state_tier_c.jsonl
print("Loading original confidences from state_tier_c.jsonl...")
orig_conf = {}
with open(STATE_LOG, "r", encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if not line:
            continue
        try:
            r = json.loads(line)
        except Exception:
            continue
        key = (r.get("university", ""), r.get("department", ""))
        orig_conf[key] = {
            "confidence": r.get("confidence"),
            "verification_status": r.get("verification_status"),
        }
print(f"  loaded {len(orig_conf)} unique (uni, dept) entries")


def dept_absent(uni, dept):
    info = orig_conf.get((uni, dept))
    if not info:
        return False
    return info["confidence"] == "not_found"


# Walk grounding files
print("Classifying...")
records = []
files = sorted(f for f in os.listdir(GROUNDING) if f.endswith(".json"))
for fn in files:
    fp = os.path.join(GROUNDING, fn)
    try:
        with open(fp, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        continue
    uni = data.get("university", "")
    dept = data.get("department", "")
    od = tier.get(uni, "") or data.get("university_domain", "")
    chunks = (data.get("grounding_metadata") or {}).get("groundingChunks") or []

    n_chunks = len(chunks)
    n_official = 0
    n_official_working = 0
    working_url = ""

    for c in chunks:
        web = (c or {}).get("web") or {}
        uri = web.get("uri", "")
        unwrap_info = unwrap.get(uri) or {}
        real_url = unwrap_info.get("real_url", "")
        if not real_url:
            chunk_domain = web.get("domain", "")
            if is_subdomain_match(chunk_domain, od):
                n_official += 1
            continue

        if is_subdomain_match(real_url, od):
            n_official += 1
            probe_info = probe.get(real_url) or {}
            if probe_info.get("ok"):
                n_official_working += 1
                if not working_url:
                    working_url = real_url

    if n_official_working >= 1:
        status, reason = "FOUND", ""
    elif n_chunks == 0:
        status, reason = "NOT_FOUND", "no grounding chunks"
    elif n_official == 0:
        status, reason = "NOT_FOUND", "chunks present but none on official domain"
    else:
        status, reason = "NOT_FOUND", f"{n_official} official chunk(s) but none probed working"

    records.append({
        "uni": uni, "dept": dept, "status": status,
        "n_chunks": n_chunks, "n_official": n_official,
        "n_official_working": n_official_working,
        "working_url": working_url, "reason": reason,
    })

total = len(records)
total_found = sum(1 for r in records if r["status"] == "FOUND")
total_nf = total - total_found
print(f"  records: {total}")
print(f"  FOUND:     {total_found}  ({total_found/total*100:.1f}%)")
print(f"  NOT FOUND: {total_nf}  ({total_nf/total*100:.1f}%)")

# Aggregate per dept
dept_stats = defaultdict(lambda: {"total": 0, "FOUND": 0, "NOT_FOUND": 0, "absent": 0})
for r in records:
    s = dept_stats[r["dept"]]
    s["total"] += 1
    s[r["status"]] += 1
    if dept_absent(r["uni"], r["dept"]):
        s["absent"] += 1

a_rows = []
for dept, s in dept_stats.items():
    cov = (s["FOUND"] / s["total"] * 100.0) if s["total"] else 0.0
    a_rows.append({
        "Department": dept,
        "Tier C universities": s["total"],
        "FOUND (official + working)": s["FOUND"],
        "NOT FOUND": s["NOT_FOUND"],
        "Coverage %": round(cov, 1),
        "Dept absent at (# unis)": s["absent"],
    })
a_rows.sort(key=lambda r: r["Coverage %"])

dept_order = {row["Department"]: i for i, row in enumerate(a_rows)}
c_rows = [r for r in records if r["status"] == "NOT_FOUND"]
c_rows.sort(key=lambda r: (dept_order.get(r["dept"], 999), r["uni"]))

# Excel
print(f"Writing {OUT_XLSX}...")
wb = Workbook()
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="305496")
red = PatternFill("solid", fgColor="F8CBAD")
amber = PatternFill("solid", fgColor="FFE699")
green = PatternFill("solid", fgColor="C6E0B4")

# Summary
ws_sum = wb.active
ws_sum.title = "Summary"
ws_sum.append(["Coverage v3 — Tier C (per-chunk HTTP probed)"])
ws_sum["A1"].font = Font(bold=True, size=14)
ws_sum.append([])
ws_sum.append(["Files processed", len(files)])
ws_sum.append(["Total (uni, dept) records", total])
ws_sum.append(["FOUND (>=1 chunk official + working)", total_found, f"{total_found/total*100:.1f}%"])
ws_sum.append(["NOT FOUND", total_nf, f"{total_nf/total*100:.1f}%"])
ws_sum.append([])
ws_sum.append(["Stage 1 (unwrap)", f"{len(unwrap)} entries"])
ws_sum.append(["Stage 2 (probe)", f"{len(probe)} entries"])
ws_sum.append([])
ws_sum.append(["FOUND criterion:"])
ws_sum.append(["", "1. Chunk's real (unwrapped) URL is on the university's official domain"])
ws_sum.append(["", "2. That real URL probed with status < 400"])
ws_sum.column_dimensions["A"].width = 50
ws_sum.column_dimensions["B"].width = 20
ws_sum.column_dimensions["C"].width = 12

# Coverage by Department
ws1 = wb.create_sheet("Coverage by Department")
headers = ["Department", "Tier C universities", "FOUND (official + working)",
           "NOT FOUND", "Coverage %", "Dept absent at (# unis)"]
ws1.append(headers)
for i in range(1, len(headers) + 1):
    c = ws1.cell(row=1, column=i)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[1].height = 36

for r in a_rows:
    ws1.append([r[h] for h in headers])
    last = ws1.max_row
    cov = r["Coverage %"]
    fill = red if cov < 50 else amber if cov < 80 else green
    ws1.cell(row=last, column=5).fill = fill

for i, w in enumerate([40, 18, 22, 14, 12, 22], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = ws1.dimensions

# Missing Universities drill-down
ws2 = wb.create_sheet("Missing Universities")
headers_c = ["Department", "University", "Dept exists at uni?",
             "Total chunks", "Official chunks", "Official + working",
             "Reason", "Best URL we have"]
ws2.append(headers_c)
for i in range(1, len(headers_c) + 1):
    c = ws2.cell(row=1, column=i)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
for r in c_rows:
    exists = "no" if dept_absent(r["uni"], r["dept"]) else "yes"
    ws2.append([
        r["dept"], r["uni"], exists,
        r["n_chunks"], r["n_official"], r["n_official_working"],
        r["reason"], r["working_url"],
    ])
for i, w in enumerate([35, 45, 18, 12, 14, 16, 50, 70], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions

wb.save(OUT_XLSX)
print(f"Done. {len(a_rows)} departments, {len(c_rows)} NOT FOUND rows.")
