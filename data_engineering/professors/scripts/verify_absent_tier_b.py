"""Cross-check the 'Dept absent at' counts using THREE independent sources:
  A) state_tier_b.jsonl  -> confidence == 'not_found'
  B) state_tier_b.jsonl  -> verification_status == 'no_match'
  C) output_tier_b/universities/  -> outcome == 'no_match'
  D) grounding_tier_b/   -> discovery_text_raw contains absence phrases

Then show which (uni, dept) pairs are flagged absent BUT v3 says FOUND.
"""
import json
import os
import re
from collections import Counter, defaultdict

BASE = r"c:\Users\HP\OneDrive\Desktop\course_data\Professors_info"

# Source A: state log confidence
state_conf = {}
state_verif = {}
with open(os.path.join(BASE, "logs", "state_tier_b.jsonl"), "r", encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if not line:
            continue
        try:
            r = json.loads(line)
        except Exception:
            continue
        key = (r.get("university", ""), r.get("department", ""))
        state_conf[key] = r.get("confidence")
        state_verif[key] = r.get("verification_status")

# Source C: output JSON outcome
output_outcome = {}
output_dir = os.path.join(BASE, "output_tier_b", "universities")
for fn in os.listdir(output_dir):
    if not fn.endswith(".json"):
        continue
    with open(os.path.join(output_dir, fn), "r", encoding="utf-8") as f:
        d = json.load(f)
    uni = d.get("university", "")
    for dept, e in d.get("departments", {}).items():
        output_outcome[(uni, dept)] = e.get("outcome")

print(f"State log entries:        {len(state_conf)}")
print(f"Output JSON entries:      {len(output_outcome)}")
print()

# ---- Counts per source ----
abs_A = sum(1 for v in state_conf.values() if v == "not_found")
abs_B = sum(1 for v in state_verif.values() if v == "no_match")
abs_C = sum(1 for v in output_outcome.values() if v == "no_match")
print("Counts of 'department absent' from each source:")
print(f"  A) state.confidence == 'not_found':       {abs_A}")
print(f"  B) state.verification_status == 'no_match': {abs_B}")
print(f"  C) output.outcome == 'no_match':           {abs_C}")
print()

# ---- Agreement ----
keys_all = set(state_conf.keys()) | set(output_outcome.keys())
agree_all = 0
A_only = []
C_only = []
both_agree = []
for k in keys_all:
    a = state_conf.get(k) == "not_found"
    c = output_outcome.get(k) == "no_match"
    if a and c:
        both_agree.append(k)
        agree_all += 1
    elif a and not c:
        A_only.append(k)
    elif c and not a:
        C_only.append(k)
print(f"Pairs where ALL THREE sources agree absent (state=not_found AND output=no_match): {agree_all}")
print(f"  In state-A but NOT output-C:  {len(A_only)}")
print(f"  In output-C but NOT state-A:  {len(C_only)}")
print()

if A_only[:5]:
    print("  Sample state-A only (state says not_found, output says ok):")
    for k in A_only[:5]:
        print(f"    {k[0]} / {k[1]}  state.verif={state_verif.get(k)}  output.outcome={output_outcome.get(k)}")
print()

# ---- Cross with v3 results ----
print("=" * 80)
print("Now cross with v3 FOUND/NOT_FOUND classifications")
print("=" * 80)
from openpyxl import load_workbook
wb = load_workbook(os.path.join(BASE, "coverage_grounding_v3_tier_b.xlsx"))
miss_ws = wb["Missing Universities"]
miss_keys = set()
for r in list(miss_ws.iter_rows(values_only=True))[1:]:
    dept, uni = r[0], r[1]
    miss_keys.add((uni, dept))

# Walk every grounding file to enumerate all v3 records
v3_records = {}
for fn in sorted(os.listdir(os.path.join(BASE, "grounding_tier_b"))):
    if not fn.endswith(".json"):
        continue
    try:
        with open(os.path.join(BASE, "grounding_tier_b", fn), "r", encoding="utf-8") as f:
            d = json.load(f)
    except Exception:
        continue
    uni = d.get("university", "")
    dept = d.get("department", "")
    v3_records[(uni, dept)] = "NOT_FOUND" if (uni, dept) in miss_keys else "FOUND"

print(f"v3 records: {len(v3_records)}")
v3_found = sum(1 for v in v3_records.values() if v == "FOUND")
v3_nf = sum(1 for v in v3_records.values() if v == "NOT_FOUND")
print(f"  FOUND: {v3_found}, NOT_FOUND: {v3_nf}")
print()

# The big question: how many records are flagged absent (any source) but v3=FOUND?
absent_keys_A = {k for k, v in state_conf.items() if v == "not_found"}
disputed = [k for k in absent_keys_A if v3_records.get(k) == "FOUND"]
print(f"Records where state-A says 'not_found' BUT v3 says FOUND: {len(disputed)}")

# Same with output source
absent_keys_C = {k for k, v in output_outcome.items() if v == "no_match"}
disputed_C = [k for k in absent_keys_C if v3_records.get(k) == "FOUND"]
print(f"Records where output-C says 'no_match' BUT v3 says FOUND:  {len(disputed_C)}")
print()

# Sample 10 disputed cases with their working URLs
print("=" * 80)
print("Sample 10 disputed (state says not_found, v3 says FOUND) — with URLs")
print("=" * 80)
# Build a uni-dept -> working_url map from the v3 Excel "Coverage" + missing
# Actually we need it from the grounding files + caches. Easier: just look at the
# stage3 "best URL" for FOUND records. But we don't store that in Excel for FOUND
# records. Let me re-derive it from the caches.
with open(os.path.join(BASE, "unwrap_cache_tier_b.json"), "r", encoding="utf-8") as f:
    unwrap = json.load(f)
with open(os.path.join(BASE, "probe_cache_tier_b.json"), "r", encoding="utf-8") as f:
    probe = json.load(f)

import csv
from urllib.parse import urlparse
tier = {}
with open(os.path.join(BASE, "config", "universities_tier_B.csv"), "r", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        tier[row["university_name"].strip()] = row.get("official_website", "").strip()

def domain_of(u):
    if not u: return ""
    s = str(u).lower()
    if "://" in s:
        try: s = urlparse(s).hostname or ""
        except: return ""
    if s.startswith("www."): s = s[4:]
    return s.strip(".")

def is_subdomain(cd, od):
    cd, od = domain_of(cd), domain_of(od)
    if not cd or not od: return False
    return cd == od or cd.endswith("." + od)

def best_url_and_chunks(uni, dept):
    fn = uni.replace(" ", "_") + "__" + re.sub(r"[ &]+", "_", dept).strip("_") + ".json"
    fp = os.path.join(BASE, "grounding_tier_b", fn)
    if not os.path.exists(fp):
        # Try harder: scan any file matching uni
        for f in os.listdir(os.path.join(BASE, "grounding_tier_b")):
            if f.startswith(uni.replace(" ", "_")):
                with open(os.path.join(BASE, "grounding_tier_b", f), "r", encoding="utf-8") as fh:
                    d = json.load(fh)
                if d.get("university") == uni and d.get("department") == dept:
                    fp = os.path.join(BASE, "grounding_tier_b", f)
                    break
    if not os.path.exists(fp):
        return ("", "", 0, "")
    with open(fp, "r", encoding="utf-8") as f:
        data = json.load(f)
    od = tier.get(uni, "") or data.get("university_domain", "")
    chunks = (data.get("grounding_metadata") or {}).get("groundingChunks") or []
    discovery = (data.get("discovery_text_raw") or "")[:200]
    for c in chunks:
        web = (c or {}).get("web") or {}
        uri = web.get("uri", "")
        ru = (unwrap.get(uri) or {}).get("real_url", "")
        if ru and is_subdomain(ru, od) and (probe.get(ru) or {}).get("ok"):
            return (ru, discovery, len(chunks), od)
    return ("", discovery, len(chunks), od)


for k in disputed[:10]:
    uni, dept = k
    url, discovery, n_ch, od = best_url_and_chunks(uni, dept)
    print(f"\n  {uni} / {dept}")
    print(f"    state.confidence: not_found, state.verif: {state_verif.get(k)}")
    print(f"    output.outcome:   {output_outcome.get(k)}")
    print(f"    chunks: {n_ch}, official_domain: {od}")
    print(f"    v3 best URL: {url}")
    print(f"    pipeline reasoning (first 200 chars):")
    print(f"      {discovery!r}")
