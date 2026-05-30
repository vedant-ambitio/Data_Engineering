from openpyxl import load_workbook
wb = load_workbook(r"c:\Users\HP\OneDrive\Desktop\course_data\Professors_info\coverage_grounding_tierA.xlsx")
ws = wb["Coverage by Department"]
rows = list(ws.iter_rows(values_only=True))
print(f"{'Department':<38} {'Total':>6} {'FOUND':>6} {'NOT_F':>6} {'Cov%':>7}")
print("-" * 70)
for r in rows[1:]:
    dept, total, found, nf, cov = r
    print(f"{dept:<38} {total:>6} {found:>6} {nf:>6} {cov:>7}")
