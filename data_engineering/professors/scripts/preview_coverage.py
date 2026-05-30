"""Quick text preview of the coverage table so we can see the answer at a glance."""
from openpyxl import load_workbook

wb = load_workbook(r"c:\Users\HP\OneDrive\Desktop\course_data\Professors_info\coverage_report.xlsx")
ws = wb["Coverage by Department"]

rows = list(ws.iter_rows(values_only=True))
header = rows[0]
data = rows[1:]

print(f"{'Department':<38} {'Off':>5} {'HIGH':>5} {'LOW':>5} {'NF':>4} {'Cov%':>6} {'A%':>6} {'BC%':>6}")
print("-" * 90)
for r in data:
    dept, offering, high, low, nf, cov, a_cov, bc_cov, total = r
    print(f"{dept:<38} {offering:>5} {high:>5} {low:>5} {nf:>4} {cov:>6} {a_cov:>6} {bc_cov:>6}")
