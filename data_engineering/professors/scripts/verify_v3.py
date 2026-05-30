"""Verify the v3 report:
   1. Show the full Coverage by Department table with the new column
   2. Spot-check 10 random Missing Universities rows
"""
import json
import os
import random
from openpyxl import load_workbook

BASE = r"c:\Users\HP\OneDrive\Desktop\course_data\Professors_info"

wb = load_workbook(os.path.join(BASE, "coverage_grounding_v3.xlsx"))

# 1. Coverage by Department
print("=" * 90)
print("Coverage by Department (full table)")
print("=" * 90)
ws = wb["Coverage by Department"]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
print(f"{'Department':<38} {'Tot':>4} {'FND':>4} {'NF':>4} {'Cov%':>6} {'Absent':>7}")
print("-" * 75)
for r in rows[1:]:
    dept, tot, fnd, nf, cov, absent = r
    print(f"{dept:<38} {tot:>4} {fnd:>4} {nf:>4} {cov:>6} {absent:>7}")

# 2. Spot-check Missing Universities
print()
print("=" * 90)
print("Spot-check: 10 random Missing Universities rows")
print("=" * 90)
ws2 = wb["Missing Universities"]
miss_rows = list(ws2.iter_rows(values_only=True))
header2 = miss_rows[0]
data2 = miss_rows[1:]
random.seed(42)
sample = random.sample(data2, 10)

# Load original data for cross-checking
output_dir = os.path.join(BASE, "output", "universities")
orig_data = {}
for fn in os.listdir(output_dir):
    if not fn.endswith(".json"):
        continue
    with open(os.path.join(output_dir, fn), "r", encoding="utf-8") as f:
        d = json.load(f)
    uni = d.get("university", "")
    for dept, e in d.get("departments", {}).items():
        orig_data[(uni, dept)] = e

print(f"{'#':>2} {'Dept':<25} {'Uni':<35} {'Exists':>7} {'OrigConf':>12} {'TotCh':>5} {'OffCh':>5} {'Reason'}")
print("-" * 130)
for i, row in enumerate(sample, 1):
    dept, uni, exists, tot_ch, off_ch, off_work, reason, url = row
    orig = orig_data.get((uni, dept), {})
    orig_conf = orig.get("confidence", "?")
    print(f"{i:>2} {dept[:25]:<25} {uni[:35]:<35} {exists:>7} {orig_conf:>12} {tot_ch:>5} {off_ch:>5} {reason}")

# 3. Sanity: count "exists=no" in Missing Universities, should match sum of "absent" col
print()
print("=" * 90)
print("Sanity check")
print("=" * 90)
exists_no = sum(1 for r in data2 if r[2] == "no")
exists_yes = sum(1 for r in data2 if r[2] == "yes")
absent_sum = sum(r[5] for r in rows[1:])  # column 5 = "Dept absent at (# unis)"
print(f"Total Missing Universities rows:               {len(data2)}")
print(f"  with 'Dept exists at uni? = no':             {exists_no}")
print(f"  with 'Dept exists at uni? = yes':            {exists_yes}")
print(f"Sum of 'Dept absent at' across all depts:      {absent_sum}")
print(f"  (should equal exists=no count)               {'✓ match' if absent_sum == exists_no else '✗ MISMATCH'}")
