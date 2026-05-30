"""Coverage analysis based on GROUNDING CHUNKS (Tier A only) — v2.

Now considers BOTH passes:
  Pass 1: grounding/{Uni}__{Dept}.json   (original Gemini grounding)
  Pass 2: output_recovery/details/{Uni}__{Dept}.json  (recovery pass)

A (uni, dept) is FOUND if EITHER pass produced:
  - At least one grounding chunk on the university's official domain, AND
  - A working verified URL (verification_status in {ok, ok_via_alternate})

Otherwise NOT FOUND.
"""
import csv
import json
import os
from collections import defaultdict
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = r"c:\Users\HP\OneDrive\Desktop\course_data\Professors_info"
OUTPUT_UNIS = os.path.join(BASE, "output", "universities")
GROUNDING_DIR = os.path.join(BASE, "grounding")
RECOVERY_DETAILS_DIR = os.path.join(BASE, "output_recovery", "details")
TIER_A_CSV = os.path.join(BASE, "config", "universities_tier_A.csv")
OUT_XLSX = os.path.join(BASE, "coverage_grounding_tierA_v2.xlsx")


# Tier A
tier_a = {}
with open(TIER_A_CSV, "r", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        name = row["university_name"].strip()
        site = row.get("official_website", "").strip()
        if name:
            tier_a[name] = site


def to_domain(url_or_host):
    if not url_or_host:
        return ""
    s = str(url_or_host).strip().lower()
    if "://" in s:
        try:
            s = urlparse(s).hostname or ""
        except Exception:
            return ""
    if s.startswith("www."):
        s = s[4:]
    return s.strip(".")


def is_subdomain_match(chunk_domain_or_url, official_domain):
    cd = to_domain(chunk_domain_or_url)
    od = to_domain(official_domain)
    if not cd or not od:
        return False
    return cd == od or cd.endswith("." + od)


def analyze_pass1(grounding_path, official_domain):
    """Original grounding/ folder."""
    if not grounding_path or not os.path.exists(grounding_path):
        return {"found": False, "has_official_chunk": False, "verified": False,
                "url": "", "reason": "no grounding file"}
    try:
        with open(grounding_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        return {"found": False, "has_official_chunk": False, "verified": False,
                "url": "", "reason": f"parse error: {e}"}

    od = official_domain or data.get("university_domain", "")
    chunks = (data.get("grounding_metadata") or {}).get("groundingChunks") or []
    has_official = False
    for c in chunks:
        web = (c or {}).get("web") or {}
        cd = web.get("domain") or web.get("uri", "")
        if is_subdomain_match(cd, od):
            has_official = True
            break

    uv = data.get("url_verification") or {}
    vstatus = uv.get("verification_status")
    verified = vstatus in ("ok", "ok_via_alternate")
    url = uv.get("verified_url") or "" if verified else ""

    if has_official and verified:
        return {"found": True, "has_official_chunk": True, "verified": True,
                "url": url, "reason": ""}
    if not chunks:
        reason = "no grounding chunks"
    elif not has_official:
        reason = "chunks present but none on official domain"
    elif not verified:
        reason = f"chunks ok, URL not verified (status={vstatus})"
    else:
        reason = "unknown"
    return {"found": False, "has_official_chunk": has_official, "verified": verified,
            "url": url, "reason": reason}


def analyze_pass2(recovery_path, official_domain):
    """Recovery pass: output_recovery/details/."""
    if not recovery_path or not os.path.exists(recovery_path):
        return {"found": False, "has_official_chunk": False, "verified": False,
                "url": "", "reason": "no recovery file"}
    try:
        with open(recovery_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        return {"found": False, "has_official_chunk": False, "verified": False,
                "url": "", "reason": f"parse error: {e}"}

    od = official_domain
    chunks = data.get("chunks") or []
    has_official = False
    for c in chunks:
        real_url = (c or {}).get("real_url", "")
        if is_subdomain_match(real_url, od):
            has_official = True
            break

    vstatus = data.get("verification_status")
    verified = vstatus in ("ok", "ok_via_alternate")
    url = data.get("final_url") or data.get("chosen_url") or "" if verified else ""

    if has_official and verified:
        return {"found": True, "has_official_chunk": True, "verified": True,
                "url": url, "reason": ""}
    return {"found": False, "has_official_chunk": has_official, "verified": verified,
            "url": url, "reason": f"recovery: chunks={'yes' if has_official else 'no'} verified={'yes' if verified else 'no'}"}


# Walk every Tier A uni
records = []
unis_processed = set()
pass1_found = pass2_rescued = 0

for fn in sorted(os.listdir(OUTPUT_UNIS)):
    if not fn.endswith(".json"):
        continue
    fp = os.path.join(OUTPUT_UNIS, fn)
    with open(fp, "r", encoding="utf-8") as f:
        data = json.load(f)
    uni = data.get("university", "")
    if uni not in tier_a:
        continue
    unis_processed.add(uni)
    od = tier_a[uni] or data.get("university_domain", "")

    for dept, entry in data.get("departments", {}).items():
        gfile = entry.get("grounding_file") or ""
        gpath = os.path.join(GROUNDING_DIR, gfile) if gfile else None
        rpath = os.path.join(RECOVERY_DETAILS_DIR, gfile) if gfile else None

        r1 = analyze_pass1(gpath, od)
        if r1["found"]:
            pass1_found += 1
            records.append({
                "university": uni, "department": dept,
                "status": "FOUND", "via": "pass1",
                "has_official_chunk": True, "verified": True,
                "url": r1["url"], "reason": ""
            })
            continue

        r2 = analyze_pass2(rpath, od)
        if r2["found"]:
            pass2_rescued += 1
            records.append({
                "university": uni, "department": dept,
                "status": "FOUND", "via": "pass2_recovery",
                "has_official_chunk": True, "verified": True,
                "url": r2["url"], "reason": ""
            })
            continue

        # Combine reasons from both passes for diagnostic clarity
        combined_reason = f"P1: {r1['reason']}"
        if os.path.exists(rpath) if rpath else False:
            combined_reason += f"; P2: {r2['reason']}"
        else:
            combined_reason += "; P2: not attempted"

        records.append({
            "university": uni, "department": dept,
            "status": "NOT_FOUND", "via": "",
            "has_official_chunk": r1["has_official_chunk"] or r2["has_official_chunk"],
            "verified": r1["verified"] or r2["verified"],
            "url": r1["url"] or r2["url"],
            "reason": combined_reason
        })

total = len(records)
total_found = sum(1 for r in records if r["status"] == "FOUND")
total_nf = total - total_found

print(f"Tier A unis processed:       {len(unis_processed)}")
print(f"Total (uni, dept) records:   {total}")
print(f"  FOUND in pass 1 grounding: {pass1_found}")
print(f"  FOUND in pass 2 recovery:  {pass2_rescued}")
print(f"  TOTAL FOUND:               {total_found}  ({total_found/total*100:.1f}%)")
print(f"  NOT FOUND:                 {total_nf}  ({total_nf/total*100:.1f}%)")


# Per-department aggregation
dept_stats = defaultdict(lambda: {"total": 0, "FOUND": 0, "NOT_FOUND": 0,
                                   "p1_found": 0, "p2_found": 0})
for r in records:
    s = dept_stats[r["department"]]
    s["total"] += 1
    s[r["status"]] += 1
    if r["via"] == "pass1":
        s["p1_found"] += 1
    elif r["via"] == "pass2_recovery":
        s["p2_found"] += 1


a_rows = []
for dept, s in dept_stats.items():
    cov = (s["FOUND"] / s["total"] * 100.0) if s["total"] else 0.0
    a_rows.append({
        "Department": dept,
        "Tier A universities": s["total"],
        "FOUND (official + working)": s["FOUND"],
        "NOT FOUND": s["NOT_FOUND"],
        "Coverage %": round(cov, 1),
    })
a_rows.sort(key=lambda r: r["Coverage %"])

dept_order = {row["Department"]: i for i, row in enumerate(a_rows)}
c_rows = [r for r in records if r["status"] == "NOT_FOUND"]
c_rows.sort(key=lambda r: (dept_order.get(r["department"], 999), r["university"]))


# Excel
wb = Workbook()
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="305496")
red = PatternFill("solid", fgColor="F8CBAD")
amber = PatternFill("solid", fgColor="FFE699")
green = PatternFill("solid", fgColor="C6E0B4")

# Summary
ws_sum = wb.active
ws_sum.title = "Summary"
ws_sum.append(["Coverage by Grounding — Tier A (pass 1 + pass 2)"])
ws_sum["A1"].font = Font(bold=True, size=14)
ws_sum.append([])
ws_sum.append(["Tier A universities processed", len(unis_processed)])
ws_sum.append(["Total (uni, dept) records", total])
ws_sum.append(["FOUND via pass 1 (grounding/)", pass1_found, f"{pass1_found/total*100:.1f}%"])
ws_sum.append(["FOUND via pass 2 (output_recovery/details/)", pass2_rescued, f"{pass2_rescued/total*100:.1f}%"])
ws_sum.append(["TOTAL FOUND", total_found, f"{total_found/total*100:.1f}%"])
ws_sum.append(["NOT FOUND", total_nf, f"{total_nf/total*100:.1f}%"])
ws_sum.append([])
ws_sum.append(["FOUND criterion (per pass):"])
ws_sum.append(["", "1. At least one grounding chunk on the official domain"])
ws_sum.append(["", "2. URL verified (status 'ok' or 'ok_via_alternate')"])
ws_sum.append(["", "Either pass succeeding counts the (uni, dept) as FOUND"])
ws_sum.column_dimensions["A"].width = 50
ws_sum.column_dimensions["B"].width = 18
ws_sum.column_dimensions["C"].width = 12

# Coverage by Department
ws1 = wb.create_sheet("Coverage by Department")
headers_a = list(a_rows[0].keys())
ws1.append(headers_a)
for i in range(1, len(headers_a) + 1):
    c = ws1.cell(row=1, column=i)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[1].height = 36

for r in a_rows:
    ws1.append([r[h] for h in headers_a])
    last = ws1.max_row
    cov = r["Coverage %"]
    fill = red if cov < 50 else amber if cov < 80 else green
    ws1.cell(row=last, column=len(headers_a)).fill = fill

for i, w in enumerate([40, 18, 22, 14, 12], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = ws1.dimensions

# Missing Universities
ws2 = wb.create_sheet("Missing Universities")
headers_c = ["Department", "University", "Has official-domain chunk?",
             "URL verified?", "Reason (P1 = pass 1, P2 = pass 2)", "Best URL we have"]
ws2.append(headers_c)
for i in range(1, len(headers_c) + 1):
    c = ws2.cell(row=1, column=i)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
for r in c_rows:
    ws2.append([
        r["department"], r["university"],
        "yes" if r["has_official_chunk"] else "no",
        "yes" if r["verified"] else "no",
        r["reason"], r["url"],
    ])
for i, w in enumerate([35, 45, 18, 14, 60, 70], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions

wb.save(OUT_XLSX)
print(f"\nWrote {OUT_XLSX}")
print(f"  - {len(a_rows)} departments")
print(f"  - {len(c_rows)} NOT FOUND rows")
