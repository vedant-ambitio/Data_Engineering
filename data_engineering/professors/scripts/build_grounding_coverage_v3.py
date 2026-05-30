"""Coverage v3 — walks every JSON in grounding/ and probes each chunk URL.

Logic per (uni, dept) grounding file:
  - For each chunk in groundingChunks, check:
      (a) chunk's domain matches the university's official domain (subdomain match)
      (b) HTTP probe of the chunk URI (after following redirects) returns status < 400
  - If at least 1 chunk satisfies BOTH (a) and (b) -> FOUND
  - Else -> NOT FOUND

Empty groundingChunks -> NOT FOUND.
"""
import csv
import json
import os
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = r"c:\Users\HP\OneDrive\Desktop\course_data\Professors_info"
GROUNDING_DIR = os.path.join(BASE, "grounding")
TIER_A_CSV = os.path.join(BASE, "config", "universities_tier_A.csv")
PROBE_CACHE = os.path.join(BASE, "probe_cache_v3.json")
OUT_XLSX = os.path.join(BASE, "coverage_grounding_v3.xlsx")

CONCURRENCY = 150
TIMEOUT = 6
USER_AGENT = "Mozilla/5.0 (compatible; CoverageBot/1.0)"

# Tier A: university_name -> official_website
tier_a = {}
with open(TIER_A_CSV, "r", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        name = row["university_name"].strip()
        site = row.get("official_website", "").strip()
        if name:
            tier_a[name] = site


def to_domain(url_or_host):
    if not url_or_host:
        return ""
    s = str(url_or_host).strip().lower()
    if "://" in s:
        try:
            s = urlparse(s).hostname or ""
        except Exception:
            return ""
    if s.startswith("www."):
        s = s[4:]
    return s.strip(".")


def is_subdomain_match(chunk_domain, official_domain):
    cd = to_domain(chunk_domain)
    od = to_domain(official_domain)
    if not cd or not od:
        return False
    return cd == od or cd.endswith("." + od)


# -------------------- Pass 1: read all grounding files --------------------
print("Pass 1: scanning grounding/ ...")
file_data = []
files = sorted(f for f in os.listdir(GROUNDING_DIR) if f.endswith(".json"))
for fn in files:
    fp = os.path.join(GROUNDING_DIR, fn)
    try:
        with open(fp, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        print(f"  parse error {fn}: {e}")
        continue
    uni = data.get("university", "")
    dept = data.get("department", "")
    od = tier_a.get(uni, "") or data.get("university_domain", "")
    chunks = (data.get("grounding_metadata") or {}).get("groundingChunks") or []
    chunk_info = []
    for c in chunks:
        web = (c or {}).get("web") or {}
        uri = web.get("uri", "")
        cd = web.get("domain") or to_domain(uri)
        chunk_info.append({
            "uri": uri,
            "domain": cd,
            "official_match": is_subdomain_match(cd, od),
        })
    file_data.append({
        "file": fn, "uni": uni, "dept": dept,
        "official_domain": od, "chunks": chunk_info,
    })

print(f"  {len(file_data)} files read")
total_chunks = sum(len(fd["chunks"]) for fd in file_data)
official_chunks = sum(sum(1 for c in fd["chunks"] if c["official_match"]) for fd in file_data)
print(f"  total chunks: {total_chunks}")
print(f"  chunks on official domain: {official_chunks}")


# -------------------- Pass 2: probe URLs (only official-domain chunks) --------------------
urls_to_probe = {c["uri"] for fd in file_data for c in fd["chunks"]
                 if c["official_match"] and c["uri"]}
print(f"\nPass 2: unique official-domain URLs to probe: {len(urls_to_probe)}")

cache = {}
if os.path.exists(PROBE_CACHE):
    try:
        with open(PROBE_CACHE, "r", encoding="utf-8") as f:
            cache = json.load(f)
        print(f"  loaded cache: {len(cache)} entries")
    except Exception as e:
        print(f"  cache load failed: {e}")
        cache = {}

unprobed = [u for u in urls_to_probe if u not in cache]
print(f"  to probe (cache miss): {len(unprobed)}")


def probe(url):
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": USER_AGENT})
        r = s.head(url, allow_redirects=True, timeout=TIMEOUT)
        if r.status_code in (400, 403, 405, 406, 501) or r.status_code >= 500:
            r = s.get(url, allow_redirects=True, timeout=TIMEOUT, stream=True)
            r.close()
        return url, {
            "status": r.status_code,
            "ok": r.status_code < 400,
            "final_url": r.url,
        }
    except requests.exceptions.Timeout:
        return url, {"status": None, "ok": False, "error": "timeout"}
    except requests.exceptions.SSLError:
        return url, {"status": None, "ok": False, "error": "ssl"}
    except Exception as e:
        return url, {"status": None, "ok": False, "error": str(e)[:120]}


if unprobed:
    start = time.time()
    done = 0
    save_every = 500
    with ThreadPoolExecutor(max_workers=CONCURRENCY) as exe:
        futs = {exe.submit(probe, u): u for u in unprobed}
        for fut in as_completed(futs):
            url, result = fut.result()
            cache[url] = result
            done += 1
            if done % 100 == 0:
                elapsed = time.time() - start
                rate = done / elapsed if elapsed else 0
                eta = (len(unprobed) - done) / rate if rate else 0
                print(f"  {done}/{len(unprobed)}  ({rate:.1f}/s  ETA {eta:.0f}s)")
            if done % save_every == 0:
                with open(PROBE_CACHE, "w", encoding="utf-8") as f:
                    json.dump(cache, f)
    with open(PROBE_CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f)
    print(f"  probing done in {time.time()-start:.0f}s")


# -------------------- Pass 3: classify each (uni, dept) --------------------
records = []
for fd in file_data:
    official_n = sum(1 for c in fd["chunks"] if c["official_match"])
    working_official_n = 0
    working_url = ""
    for c in fd["chunks"]:
        if c["official_match"] and c["uri"]:
            res = cache.get(c["uri"]) or {}
            if res.get("ok"):
                working_official_n += 1
                if not working_url:
                    working_url = res.get("final_url") or c["uri"]

    if working_official_n >= 1:
        status = "FOUND"
        reason = ""
    else:
        if not fd["chunks"]:
            reason = "no grounding chunks"
        elif official_n == 0:
            reason = "chunks present but none on official domain"
        else:
            reason = f"{official_n} official chunk(s) but none probed working"
        status = "NOT_FOUND"

    records.append({
        "uni": fd["uni"], "dept": fd["dept"],
        "status": status,
        "official_chunks": official_n,
        "working_official_chunks": working_official_n,
        "working_url": working_url,
        "reason": reason,
    })


# -------------------- Aggregate per department --------------------
dept_stats = defaultdict(lambda: {"total": 0, "FOUND": 0, "NOT_FOUND": 0})
for r in records:
    s = dept_stats[r["dept"]]
    s["total"] += 1
    s[r["status"]] += 1

a_rows = []
for dept, s in dept_stats.items():
    cov = (s["FOUND"] / s["total"] * 100.0) if s["total"] else 0.0
    a_rows.append({
        "Department": dept,
        "Tier A universities": s["total"],
        "FOUND (official + working)": s["FOUND"],
        "NOT FOUND": s["NOT_FOUND"],
        "Coverage %": round(cov, 1),
    })
a_rows.sort(key=lambda r: r["Coverage %"])

dept_order = {row["Department"]: i for i, row in enumerate(a_rows)}
c_rows = [r for r in records if r["status"] == "NOT_FOUND"]
c_rows.sort(key=lambda r: (dept_order.get(r["dept"], 999), r["uni"]))


# -------------------- Excel --------------------
wb = Workbook()
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="305496")
red = PatternFill("solid", fgColor="F8CBAD")
amber = PatternFill("solid", fgColor="FFE699")
green = PatternFill("solid", fgColor="C6E0B4")

# Summary
ws_sum = wb.active
ws_sum.title = "Summary"
ws_sum.append(["Coverage v3 — every chunk URL HTTP-probed"])
ws_sum["A1"].font = Font(bold=True, size=14)
total = len(records)
total_found = sum(1 for r in records if r["status"] == "FOUND")
total_nf = total - total_found
ws_sum.append([])
ws_sum.append(["Files processed", len(file_data)])
ws_sum.append(["Total (uni, dept) records", total])
ws_sum.append(["FOUND (>=1 chunk official + working)", total_found, f"{total_found/total*100:.1f}%"])
ws_sum.append(["NOT FOUND", total_nf, f"{total_nf/total*100:.1f}%"])
ws_sum.append([])
ws_sum.append(["Total chunks across all files", total_chunks])
ws_sum.append(["Chunks on official domain", official_chunks])
ws_sum.append(["Unique URLs probed", len(urls_to_probe)])
ws_sum.append([])
ws_sum.append(["FOUND criterion:"])
ws_sum.append(["", "1. At least one chunk's domain matches the university's official domain"])
ws_sum.append(["", "2. That chunk's URI returns status < 400 after following redirects"])
ws_sum.column_dimensions["A"].width = 50
ws_sum.column_dimensions["B"].width = 18
ws_sum.column_dimensions["C"].width = 12

# Coverage by Department
ws1 = wb.create_sheet("Coverage by Department")
headers = ["Department", "Tier A universities", "FOUND (official + working)",
           "NOT FOUND", "Coverage %"]
ws1.append(headers)
for i in range(1, len(headers) + 1):
    c = ws1.cell(row=1, column=i)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[1].height = 36

for r in a_rows:
    ws1.append([r[h] for h in headers])
    last = ws1.max_row
    cov = r["Coverage %"]
    fill = red if cov < 50 else amber if cov < 80 else green
    ws1.cell(row=last, column=5).fill = fill

for i, w in enumerate([40, 18, 22, 14, 12], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = ws1.dimensions

# Missing Universities (drill-down)
ws2 = wb.create_sheet("Missing Universities")
headers_c = ["Department", "University", "Official chunks",
             "Working official chunks", "Reason", "Best URL we have"]
ws2.append(headers_c)
for i in range(1, len(headers_c) + 1):
    c = ws2.cell(row=1, column=i)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
for r in c_rows:
    ws2.append([
        r["dept"], r["uni"],
        r["official_chunks"], r["working_official_chunks"],
        r["reason"], r["working_url"],
    ])
for i, w in enumerate([35, 45, 14, 20, 50, 70], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions

wb.save(OUT_XLSX)
print(f"\nWrote {OUT_XLSX}")
print(f"  FOUND: {total_found}/{total}  ({total_found/total*100:.1f}%)")
print(f"  NOT FOUND: {total_nf}")
