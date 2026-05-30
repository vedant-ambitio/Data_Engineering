"""How many NOT_FOUND records are because the department doesn't exist at the uni?
Per-department impact on coverage if we exclude these.
"""
import json
import os
from collections import defaultdict

BASE = r"c:\Users\HP\OneDrive\Desktop\course_data\Professors_info"
OUTPUT_DIR = os.path.join(BASE, "output", "universities")

# Load original confidence per (uni, dept)
orig = {}
for fn in os.listdir(OUTPUT_DIR):
    if not fn.endswith(".json"):
        continue
    with open(os.path.join(OUTPUT_DIR, fn), "r", encoding="utf-8") as f:
        data = json.load(f)
    uni = data.get("university", "")
    for dept, entry in data.get("departments", {}).items():
        orig[(uni, dept)] = {
            "confidence": entry.get("confidence"),
            "pre_filtered": entry.get("pre_filtered", False),
        }

# Walk the v3 results (rebuild from caches like stage3 does)
from openpyxl import load_workbook
wb = load_workbook(os.path.join(BASE, "coverage_grounding_v3.xlsx"))
miss_ws = wb["Missing Universities"]
nf_set = set()
for row in list(miss_ws.iter_rows(values_only=True))[1:]:
    dept, uni = row[0], row[1]
    nf_set.add((uni, dept))

# Coverage sheet for found counts
cov_ws = wb["Coverage by Department"]
cov_rows = list(cov_ws.iter_rows(values_only=True))[1:]

# Per-dept: total, NOT_FOUND, of which "department doesn't exist"
dept_total = defaultdict(int)
dept_found = defaultdict(int)
dept_dept_absent = defaultdict(int)  # NOT_FOUND because dept doesn't exist
dept_real_nf = defaultdict(int)      # NOT_FOUND for other reasons

# Iterate all (uni, dept) using orig
for (uni, dept), info in orig.items():
    dept_total[dept] += 1
    is_nf = (uni, dept) in nf_set
    is_dept_absent = (info["confidence"] == "not_found"
                      or info["confidence"] == "not_applicable"
                      or info["pre_filtered"])
    if is_nf:
        if is_dept_absent:
            dept_dept_absent[dept] += 1
        else:
            dept_real_nf[dept] += 1
    else:
        dept_found[dept] += 1

# Print table
rows = []
for dept in sorted(dept_total):
    tot = dept_total[dept]
    fnd = dept_found[dept]
    absent = dept_dept_absent[dept]
    real = dept_real_nf[dept]
    realistic_total = tot - absent  # universities that should have the dept
    cov_raw = fnd / tot * 100 if tot else 0
    cov_realistic = fnd / realistic_total * 100 if realistic_total else 0
    rows.append((dept, tot, fnd, absent, real, cov_raw, cov_realistic))

# Sort by realistic coverage
rows.sort(key=lambda r: r[6])

print(f"{'Department':<38} {'Tot':>4} {'FND':>4} {'NoDpt':>5} {'RealNF':>6} {'Raw%':>6} {'Real%':>6}")
print("-" * 80)
for dept, tot, fnd, absent, real, cov_raw, cov_realistic in rows:
    print(f"{dept:<38} {tot:>4} {fnd:>4} {absent:>5} {real:>6} {cov_raw:>6.1f} {cov_realistic:>6.1f}")

print()
total_records = sum(dept_total.values())
total_dept_absent = sum(dept_dept_absent.values())
total_real_nf = sum(dept_real_nf.values())
total_found = sum(dept_found.values())
print(f"TOTAL records:                            {total_records}")
print(f"TOTAL FOUND:                              {total_found}")
print(f"TOTAL NOT_FOUND - dept doesn't exist:     {total_dept_absent}")
print(f"TOTAL NOT_FOUND - real coverage gap:      {total_real_nf}")
print()
print(f"Raw coverage %:        {total_found/total_records*100:.1f}%")
print(f"Realistic coverage %:  {total_found/(total_records-total_dept_absent)*100:.1f}%   (excluding 'dept doesn't exist')")
