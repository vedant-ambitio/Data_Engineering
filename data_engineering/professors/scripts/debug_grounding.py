"""Diagnose why entries are NOT FOUND under the grounding-chunk criterion."""
from collections import Counter
from openpyxl import load_workbook

wb = load_workbook(r"c:\Users\HP\OneDrive\Desktop\course_data\Professors_info\coverage_grounding_tierA.xlsx")
ws = wb["Missing Universities"]

reasons = Counter()
chunk_yes_verified_no = 0
chunk_no_verified_yes = 0
chunk_no_verified_no = 0
chunk_yes_verified_yes_but_not_found = 0  # shouldn't happen

for row in ws.iter_rows(min_row=2, values_only=True):
    dept, uni, has_chunk, verified, reason, url = row
    reasons[reason] += 1
    if has_chunk == "yes" and verified == "no":
        chunk_yes_verified_no += 1
    elif has_chunk == "no" and verified == "yes":
        chunk_no_verified_yes += 1
    elif has_chunk == "no" and verified == "no":
        chunk_no_verified_no += 1
    elif has_chunk == "yes" and verified == "yes":
        chunk_yes_verified_yes_but_not_found += 1

total_nf = sum(reasons.values())
print(f"Total NOT FOUND: {total_nf}")
print()
print("Reason breakdown:")
for r, c in reasons.most_common():
    print(f"  {c:5d}  ({c/total_nf*100:5.1f}%)  {r}")
print()
print("Cross-tab:")
print(f"  chunk=yes, verified=yes (impossible)        {chunk_yes_verified_yes_but_not_found}")
print(f"  chunk=yes, verified=no                      {chunk_yes_verified_no}")
print(f"  chunk=no,  verified=yes (Google missed it)  {chunk_no_verified_yes}")
print(f"  chunk=no,  verified=no                      {chunk_no_verified_no}")
